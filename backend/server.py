from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, Request, BackgroundTasks
from fastapi.responses import HTMLResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import os
import logging
import asyncio
import io
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import stripe
import resend
import mercadopago
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import cloudinary
import cloudinary.uploader
import base64

# Import for image moderation
from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Cloudinary Configuration
cloudinary.config(
    cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key=os.environ.get('CLOUDINARY_API_KEY'),
    api_secret=os.environ.get('CLOUDINARY_API_SECRET')
)

# Emergent LLM Key for image moderation
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'achaservico')]

# Auth Backend URL (Emergent Authentication Service)
AUTH_BACKEND_URL = os.environ.get('AUTH_BACKEND_URL', 'https://demobackend.emergentagent.com')

# Resend Email Configuration
RESEND_API_KEY = os.environ.get('RESEND_API_KEY')
ADMIN_NOTIFICATION_EMAIL = os.environ.get('ADMIN_NOTIFICATION_EMAIL', 'saragomeshh@gmail.com')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')

if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# PIX Manual Configuration (from environment variables)
PIX_KEY = os.environ.get('PIX_KEY', '49958688875')
PIX_KEY_TYPE = os.environ.get('PIX_KEY_TYPE', 'cpf')
PIX_RECEIVER_NAME = os.environ.get('PIX_RECEIVER_NAME', 'AchaServico')

# Stripe Configuration
STRIPE_SECRET_KEY = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_PUBLISHABLE_KEY = os.environ.get('STRIPE_PUBLISHABLE_KEY', '')
STRIPE_WEBHOOK_SECRET = os.environ.get('STRIPE_WEBHOOK_SECRET', '')
stripe.api_key = STRIPE_SECRET_KEY

# Mercado Pago Configuration
MP_ACCESS_TOKEN = os.environ.get('MP_ACCESS_TOKEN', '')
MP_PUBLIC_KEY = os.environ.get('MP_PUBLIC_KEY', '')
mp_sdk = mercadopago.SDK(MP_ACCESS_TOKEN) if MP_ACCESS_TOKEN else None

# App Domain for Stripe redirects
# App configuration
APP_DOMAIN = os.environ.get('APP_DOMAIN', 'https://achaservico-backend.onrender.com')
# For mobile app deep linking after Stripe payment
APP_SCHEME = os.environ.get('APP_SCHEME', 'achaservico')

# Create the main app
app = FastAPI(title="AchaServiço API", description="API para conectar clientes a prestadores de serviços locais")

# ======================== RATE LIMITING ========================
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Admin email constant
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'israel.moraes03@gmail.com')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ======================== IMAGE MODERATION ========================

async def moderate_image(image_base64: str) -> dict:
    """
    Analyze image using AI to detect inappropriate content.
    Returns: {"is_appropriate": bool, "reason": str}
    """
    logger.info("Starting image moderation...")
    
    if not EMERGENT_LLM_KEY:
        logger.warning("EMERGENT_LLM_KEY not configured, allowing image")
        return {"is_appropriate": True, "reason": "Moderation not configured"}
    
    try:
        # Remove base64 prefix if present
        original_base64 = image_base64
        if "," in image_base64:
            image_base64 = image_base64.split(",")[1]
        
        # Check if image is too small (likely not a real image)
        if len(image_base64) < 100:
            logger.warning("Image too small, allowing")
            return {"is_appropriate": True, "reason": "Image too small to analyze"}
        
        logger.info(f"Moderating image, size: {len(image_base64)} chars")
        
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"moderation-{uuid.uuid4()}",
            system_message="""Você é um moderador de conteúdo. Analise a imagem.

REJEITE APENAS se a imagem contiver:
- Nudez explícita (genitais ou seios totalmente expostos)
- Pornografia
- Violência gráfica extrema

ACEITE todas as outras imagens, incluindo:
- Fotos de pessoas vestidas
- Biquíni, praia, piscina
- Fotos de trabalho/serviços
- Qualquer foto normal

Na dúvida, ACEITE a imagem.

Responda APENAS: {"is_appropriate": true} ou {"is_appropriate": false, "reason": "motivo"}"""
        ).with_model("openai", "gpt-4o-mini")
        
        image_content = ImageContent(image_base64=image_base64)
        
        user_message = UserMessage(
            text="Esta imagem é apropriada?",
            image_contents=[image_content]
        )
        
        response = await chat.send_message(user_message)
        logger.info(f"Moderation response: {response}")
        
        # Parse response
        import json
        try:
            response_text = response.strip()
            # Remove markdown code blocks if present
            if "```" in response_text:
                parts = response_text.split("```")
                for part in parts:
                    if "{" in part and "}" in part:
                        response_text = part.replace("json", "").strip()
                        break
            
            # Find JSON in response
            start = response_text.find("{")
            end = response_text.rfind("}") + 1
            if start != -1 and end > start:
                json_str = response_text[start:end]
                result = json.loads(json_str)
                logger.info(f"Moderation result: {result}")
                return result
        except Exception as parse_error:
            logger.warning(f"Could not parse JSON: {parse_error}")
        
        # If can't parse, check for explicit rejection
        response_lower = response.lower()
        if "false" in response_lower and ("nude" in response_lower or "porn" in response_lower or "explícit" in response_lower):
            return {"is_appropriate": False, "reason": "Conteúdo impróprio detectado"}
        
        # Default: allow the image
        return {"is_appropriate": True, "reason": "Approved"}
            
    except Exception as e:
        logger.error(f"Error moderating image: {e}")
        # On error, ALLOW the image (fail-open for better user experience)
        return {"is_appropriate": True, "reason": "Moderation skipped"}

# ======================== CLOUDINARY UPLOAD ========================

async def upload_to_cloudinary(base64_image: str, folder: str = "achaservico") -> Optional[str]:
    """
    Upload base64 image to Cloudinary and return the URL.
    Returns None if upload fails.
    """
    try:
        # Check if Cloudinary is configured
        if not os.environ.get('CLOUDINARY_CLOUD_NAME'):
            logger.warning("Cloudinary not configured, returning original base64")
            return base64_image
        
        # If it's already a URL, return as is
        if base64_image.startswith('http'):
            return base64_image
        
        # Upload to Cloudinary
        result = cloudinary.uploader.upload(
            base64_image,
            folder=folder,
            resource_type="image",
            transformation=[
                {"quality": "auto:good"},
                {"fetch_format": "auto"}
            ]
        )
        
        logger.info(f"Image uploaded to Cloudinary: {result.get('secure_url')}")
        return result.get('secure_url')
    except Exception as e:
        logger.error(f"Error uploading to Cloudinary: {e}")
        # Return original base64 if upload fails
        return base64_image

async def upload_images_to_cloudinary(images: List[str], folder: str = "achaservico") -> List[str]:
    """Upload multiple images to Cloudinary"""
    uploaded_urls = []
    for img in images:
        url = await upload_to_cloudinary(img, folder)
        if url:
            uploaded_urls.append(url)
    return uploaded_urls

# ======================== EMAIL NOTIFICATIONS ========================

async def send_expiration_notification_email(provider_name: str, provider_email: str, days_until_expiration: int):
    """Send email notification to provider when subscription is about to expire"""
    if not RESEND_API_KEY:
        logger.warning("RESEND_API_KEY not configured, skipping expiration email")
        return False
    
    try:
        if days_until_expiration <= 0:
            subject = f"⚠️ {provider_name}, sua assinatura venceu!"
            message = "Sua assinatura no AchaServiço venceu. Renove agora para continuar recebendo clientes!"
            urgency_color = "#dc3545"
        else:
            subject = f"⏰ {provider_name}, sua assinatura vence em {days_until_expiration} dia(s)!"
            message = f"Sua assinatura no AchaServiço vence em {days_until_expiration} dia(s). Renove agora e continue recebendo clientes!"
            urgency_color = "#ffc107"
        
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, {urgency_color}, #856404); padding: 20px; border-radius: 10px 10px 0 0;">
                <h1 style="color: white; margin: 0; font-size: 24px;">⏰ Aviso de Assinatura</h1>
            </div>
            <div style="background: #f9f9f9; padding: 20px; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 10px 10px;">
                <h2 style="color: #333; margin-top: 0;">Olá, {provider_name}!</h2>
                <p style="font-size: 16px; color: #333; line-height: 1.6;">
                    {message}
                </p>
                <div style="text-align: center; margin: 30px 0;">
                    <a href="achaservico://dashboard" style="background: #10B981; color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: bold; font-size: 16px;">
                        🚀 Renovar Agora
                    </a>
                </div>
                <p style="color: #666; font-size: 14px; text-align: center;">
                    Continue aparecendo para milhares de clientes em busca de serviços!
                </p>
                <p style="color: #999; font-size: 12px; margin-top: 20px; text-align: center;">
                    Este email foi enviado automaticamente pelo sistema AchaServiço.
                </p>
            </div>
        </div>
        """
        
        params = {
            "from": SENDER_EMAIL,
            "to": [provider_email],
            "subject": subject,
            "html": html_content
        }
        
        await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Expiration notification email sent to: {provider_email}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send expiration notification email: {str(e)}")
        return False

async def send_push_notification(push_token: str, title: str, body: str, data: dict = None):
    """Send push notification via Expo with proper Android background support"""
    if not push_token or not push_token.startswith("ExponentPushToken"):
        logger.warning(f"Invalid push token: {push_token[:20] if push_token else 'None'}...")
        return False
    
    try:
        notification_data = data or {"type": "general"}
        
        # Build notification payload optimized for Android background delivery
        payload = {
            "to": push_token,
            "title": title,
            "body": body,
            "sound": "default",
            "priority": "high",
            "channelId": "default",
            "data": notification_data,
            # Critical for Android background notifications
            "mutableContent": True,
            "contentAvailable": True,
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://exp.host/--/api/v2/push/send",
                json=payload,
                headers={
                    "Content-Type": "application/json",
                    "Accept": "application/json",
                    "Accept-Encoding": "gzip, deflate"
                },
                timeout=30.0
            )
            
            result = response.json()
            logger.info(f"Push API response: {result}")
            
            if response.status_code == 200:
                # Check for errors in response data
                if isinstance(result.get("data"), dict):
                    if result["data"].get("status") == "error":
                        error_msg = result["data"].get("message", "Unknown error")
                        logger.error(f"Push notification error: {error_msg}")
                        return False
                elif isinstance(result.get("data"), list):
                    for item in result["data"]:
                        if item.get("status") == "error":
                            logger.error(f"Push error: {item.get('message')}")
                            return False
                
                logger.info(f"Push notification sent successfully to: {push_token[:30]}...")
                return True
            else:
                logger.error(f"Push notification HTTP failed: {response.status_code} - {response.text}")
                return False
    except Exception as e:
        logger.error(f"Error sending push notification: {str(e)}")
        return False

async def check_expiring_subscriptions():
    """Check for subscriptions expiring in 1 day and send notifications"""
    try:
        tomorrow = datetime.now(timezone.utc) + timedelta(days=1)
        today = datetime.now(timezone.utc)
        
        # Find providers whose subscription expires in 1 day (and are not premium)
        expiring_providers = await db.providers.find({
            "subscription_status": "active",
            "is_premium": {"$ne": True},
            "subscription_expires_at": {
                "$gte": today,
                "$lte": tomorrow
            }
        }, {"_id": 0}).to_list(100)
        
        # Find providers whose subscription already expired
        expired_providers = await db.providers.find({
            "subscription_status": "active",
            "is_premium": {"$ne": True},
            "subscription_expires_at": {"$lt": today}
        }, {"_id": 0}).to_list(100)
        
        notifications_sent = 0
        
        # Send notifications for expiring subscriptions
        for provider in expiring_providers:
            user = await db.users.find_one({"user_id": provider.get("user_id")}, {"_id": 0})
            if user:
                # Calculate days until expiration
                expires_at = provider.get("subscription_expires_at")
                if expires_at:
                    days_left = (expires_at - today).days
                    
                    # Send email
                    await send_expiration_notification_email(
                        provider.get("name"),
                        user.get("email"),
                        days_left
                    )
                    
                    # Send push notification
                    push_token = provider.get("push_token")
                    if push_token:
                        await send_push_notification(
                            push_token,
                            "⏰ Assinatura expirando!",
                            f"Sua assinatura vence em {days_left} dia(s). Renove agora! 🚀"
                        )
                    
                    notifications_sent += 1
        
        # Update and notify expired subscriptions
        for provider in expired_providers:
            # Update status to expired
            await db.providers.update_one(
                {"provider_id": provider.get("provider_id")},
                {"$set": {"subscription_status": "expired", "is_active": False}}
            )
            
            user = await db.users.find_one({"user_id": provider.get("user_id")}, {"_id": 0})
            if user:
                await send_expiration_notification_email(
                    provider.get("name"),
                    user.get("email"),
                    0  # Already expired
                )
                
                push_token = provider.get("push_token")
                if push_token:
                    await send_push_notification(
                        push_token,
                        "⚠️ Assinatura vencida!",
                        "Sua assinatura venceu. Renove agora para continuar recebendo clientes!"
                    )
                
                notifications_sent += 1
        
        logger.info(f"Expiration check completed: {notifications_sent} notifications sent")
        return notifications_sent
        
    except Exception as e:
        logger.error(f"Error checking expiring subscriptions: {str(e)}")
        return 0

async def send_review_reminders():
    """Send push notifications to users 24h after contacting a provider, reminding them to leave a review"""
    try:
        now = datetime.now(timezone.utc)
        # Find contacts from 24h ago (with 1 hour window) that haven't received reminder yet
        time_24h_ago = now - timedelta(hours=24)
        time_25h_ago = now - timedelta(hours=25)
        
        contacts_to_remind = await db.whatsapp_contacts.find({
            "contacted_at": {
                "$gte": time_25h_ago,
                "$lte": time_24h_ago
            },
            "review_reminder_sent": {"$ne": True}
        }).to_list(100)
        
        reminders_sent = 0
        
        for contact in contacts_to_remind:
            user_id = contact.get("user_id")
            provider_id = contact.get("provider_id")
            
            # Get user's push token
            user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "push_token": 1, "name": 1})
            if not user or not user.get("push_token"):
                continue
            
            # Get provider name
            provider = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0, "name": 1})
            if not provider:
                continue
            
            provider_name = provider.get("name", "o prestador")
            
            # Check if user already reviewed this provider
            existing_review = await db.reviews.find_one({
                "user_id": user_id,
                "provider_id": provider_id
            })
            
            if existing_review:
                # Already reviewed, just mark as sent and skip
                await db.whatsapp_contacts.update_one(
                    {"contact_id": contact.get("contact_id")},
                    {"$set": {"review_reminder_sent": True}}
                )
                continue
            
            # Send push notification
            try:
                success = await send_push_notification(
                    push_token=user.get("push_token"),
                    title="⭐ Como foi o serviço?",
                    body=f"Você contatou {provider_name} ontem. Que tal deixar uma avaliação?",
                    data={
                        "type": "review_reminder",
                        "provider_id": provider_id
                    }
                )
                
                if success:
                    reminders_sent += 1
                    
                # Mark as sent regardless of success (to avoid spam)
                await db.whatsapp_contacts.update_one(
                    {"contact_id": contact.get("contact_id")},
                    {"$set": {"review_reminder_sent": True}}
                )
                
            except Exception as e:
                logger.error(f"Failed to send review reminder to {user_id}: {e}")
        
        logger.info(f"Review reminders sent: {reminders_sent}")
        return reminders_sent
        
    except Exception as e:
        logger.error(f"Error sending review reminders: {str(e)}")
        return 0

async def send_admin_notification_email(provider_name: str, provider_email: str, categories: List[str], cities: List[str], phone: str):
    """Send email notification to admin when a new provider registers"""
    if not RESEND_API_KEY:
        logger.warning("RESEND_API_KEY not configured, skipping email notification")
        return False
    
    try:
        categories_str = ", ".join(categories) if categories else "Não informado"
        cities_str = ", ".join([c.replace("_", " ").title() for c in cities]) if cities else "Não informado"
        
        html_content = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, #10B981, #059669); padding: 20px; border-radius: 10px 10px 0 0;">
                <h1 style="color: white; margin: 0; font-size: 24px;">🎉 Novo Prestador Cadastrado!</h1>
            </div>
            <div style="background: #f9f9f9; padding: 20px; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 10px 10px;">
                <h2 style="color: #333; margin-top: 0;">Detalhes do Cadastro:</h2>
                <table style="width: 100%; border-collapse: collapse;">
                    <tr>
                        <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; font-weight: bold; color: #666;">Nome:</td>
                        <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; color: #333;">{provider_name}</td>
                    </tr>
                    <tr>
                        <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; font-weight: bold; color: #666;">Email:</td>
                        <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; color: #333;">{provider_email}</td>
                    </tr>
                    <tr>
                        <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; font-weight: bold; color: #666;">Telefone:</td>
                        <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; color: #333;">{phone}</td>
                    </tr>
                    <tr>
                        <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; font-weight: bold; color: #666;">Categorias:</td>
                        <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; color: #333;">{categories_str}</td>
                    </tr>
                    <tr>
                        <td style="padding: 10px 0; font-weight: bold; color: #666;">Cidades:</td>
                        <td style="padding: 10px 0; color: #333;">{cities_str}</td>
                    </tr>
                </table>
                <div style="margin-top: 20px; padding: 15px; background: #fff3cd; border-radius: 5px; border-left: 4px solid #ffc107;">
                    <p style="margin: 0; color: #856404;">
                        <strong>⚠️ Ação necessária:</strong> O prestador precisa ativar a assinatura para aparecer nas buscas.
                    </p>
                </div>
                <p style="color: #666; font-size: 12px; margin-top: 20px; text-align: center;">
                    Este email foi enviado automaticamente pelo sistema AchaServiço.
                </p>
            </div>
        </div>
        """
        
        params = {
            "from": SENDER_EMAIL,
            "to": [ADMIN_NOTIFICATION_EMAIL],
            "subject": f"🆕 Novo Prestador: {provider_name}",
            "html": html_content
        }
        
        # Run sync SDK in thread to keep FastAPI non-blocking
        email_result = await asyncio.to_thread(resend.Emails.send, params)
        logger.info(f"Admin notification email sent for new provider: {provider_name}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send admin notification email: {str(e)}")
        return False

# ======================== MODELS ========================

class User(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    phone: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_provider: bool = False
    favorite_providers: List[str] = []  # List of provider_ids
    blocked: bool = False

class UserSession(BaseModel):
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Provider(BaseModel):
    provider_id: str = Field(default_factory=lambda: f"prov_{uuid.uuid4().hex[:12]}")
    user_id: str
    name: str
    phone: str
    categories: List[str] = []  # Multiple categories
    cities: List[str] = ["tres_lagoas"]  # Multiple cities of operation
    neighborhood: Optional[str] = None  # Legacy field (single)
    neighborhoods: List[str] = []  # New field (multiple)
    description: str
    profile_image: Optional[str] = None  # base64
    service_photos: List[str] = []  # Array of base64 images for service gallery
    average_rating: float = 0.0
    total_reviews: int = 0
    is_active: bool = True
    is_premium: bool = False  # Premium providers have lifetime subscription
    is_verified: bool = False  # Verified badge - 5+ positive reviews (4+ stars)
    is_available_now: bool = False  # Provider is available for immediate service
    subscription_status: str = "inactive"  # active, inactive, expired
    subscription_expires_at: Optional[datetime] = None
    mp_preference_id: Optional[str] = None
    push_token: Optional[str] = None  # Expo push notification token
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProviderCreate(BaseModel):
    name: str
    phone: str
    categories: List[str]  # Multiple categories
    cities: List[str] = ["tres_lagoas"]  # Multiple cities of operation
    neighborhood: Optional[str] = None  # Legacy field
    neighborhoods: List[str] = []  # New field (multiple)
    description: str
    profile_image: Optional[str] = None

class ProviderUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    categories: Optional[List[str]] = None  # Multiple categories
    cities: Optional[List[str]] = None  # Multiple cities of operation
    neighborhood: Optional[str] = None  # Legacy field
    neighborhoods: Optional[List[str]] = None  # New field (multiple)
    description: Optional[str] = None
    profile_image: Optional[str] = None
    service_photos: Optional[List[str]] = None  # Array of base64 images for service gallery

class Review(BaseModel):
    review_id: str = Field(default_factory=lambda: f"rev_{uuid.uuid4().hex[:12]}")
    provider_id: str
    user_id: str
    # user_name is stored but NOT returned to providers (anonymous)
    user_name: str
    rating: int  # 1-5
    comment: Optional[str] = None
    is_verified: bool = True  # Verified contact via WhatsApp
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ReviewCreate(BaseModel):
    provider_id: str
    rating: int
    comment: Optional[str] = None

# Model to track WhatsApp contacts (for verified reviews)
class WhatsAppContact(BaseModel):
    contact_id: str = Field(default_factory=lambda: f"contact_{uuid.uuid4().hex[:12]}")
    user_id: str
    provider_id: str
    contacted_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    review_reminder_sent: bool = False  # Track if 24h reminder was sent

class Notification(BaseModel):
    notification_id: str = Field(default_factory=lambda: f"notif_{uuid.uuid4().hex[:12]}")
    user_id: str  # recipient user_id
    title: str
    message: str
    is_read: bool = False
    notification_type: str = "broadcast"  # broadcast, system, personal
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Subscription(BaseModel):
    subscription_id: str = Field(default_factory=lambda: f"sub_{uuid.uuid4().hex[:12]}")
    provider_id: str
    user_id: str
    amount: float = 9.99
    status: str = "pending"  # pending, active, expired, cancelled
    payment_method: str = "mercadopago"
    mp_preference_id: Optional[str] = None
    mp_payment_id: Optional[str] = None
    started_at: Optional[datetime] = None
    expires_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Report(BaseModel):
    report_id: str = Field(default_factory=lambda: f"rep_{uuid.uuid4().hex[:12]}")
    provider_id: str
    provider_name: Optional[str] = None
    reporter_user_id: Optional[str] = None
    reporter_email: Optional[str] = None
    reason: str  # inappropriate_content, false_info, bad_behavior, other
    description: Optional[str] = None
    status: str = "pending"  # pending, accepted, discarded
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    resolved_at: Optional[datetime] = None

class ReportCreate(BaseModel):
    provider_id: str
    reason: str
    description: Optional[str] = None

class SessionDataResponse(BaseModel):
    id: str
    email: str
    name: str
    picture: Optional[str] = None
    session_token: str

# ======================== STATIC DATA ========================

CATEGORIES = [
    {"id": "eletricista", "name": "Eletricista", "icon": "flash"},
    {"id": "encanador", "name": "Encanador", "icon": "water"},
    {"id": "pedreiro", "name": "Pedreiro", "icon": "construct"},
    {"id": "pintor", "name": "Pintor", "icon": "color-palette"},
    {"id": "jardineiro", "name": "Jardineiro", "icon": "leaf"},
    {"id": "diarista", "name": "Diarista", "icon": "home"},
    {"id": "mecanico", "name": "Mecânico", "icon": "car"},
    {"id": "carpinteiro", "name": "Carpinteiro", "icon": "hammer"},
    {"id": "serralheiro", "name": "Serralheiro", "icon": "build"},
    {"id": "ar_condicionado", "name": "Técnico de Ar Condicionado", "icon": "snow"},
    {"id": "informatica", "name": "Técnico de Informática", "icon": "laptop"},
    {"id": "manicure", "name": "Manicure/Pedicure", "icon": "hand-left"},
    {"id": "cabeleireiro", "name": "Cabeleireiro(a)", "icon": "cut"},
    {"id": "design_sobrancelha", "name": "Design de Sobrancelha", "icon": "eye"},
    {"id": "personal", "name": "Personal Trainer", "icon": "fitness"},
    {"id": "professor", "name": "Professor Particular", "icon": "school"},
    {"id": "costureira", "name": "Costureira", "icon": "shirt"},
    {"id": "fotografo", "name": "Fotógrafo", "icon": "camera"},
    {"id": "montador_moveis", "name": "Montador de Móveis", "icon": "bed"},
    {"id": "taxista", "name": "Taxista", "icon": "car-sport"},
    {"id": "guincho", "name": "Guincho", "icon": "car"},
    {"id": "lava_passa", "name": "Lava e Passa Roupa", "icon": "shirt"},
    {"id": "baba", "name": "Babá", "icon": "people"},
    {"id": "baba_pet", "name": "Babá de Pet", "icon": "paw"},
    {"id": "veterinario", "name": "Veterinário Domicílio", "icon": "medkit"},
    {"id": "maquiadora", "name": "Maquiadora", "icon": "color-palette"},
    {"id": "chaveiro", "name": "Chaveiro", "icon": "key"},
    {"id": "garcom", "name": "Garçom", "icon": "restaurant"},
    {"id": "freelancer", "name": "Free-lancer", "icon": "briefcase"},
    {"id": "motoboy", "name": "Motoboy", "icon": "bicycle"},
    {"id": "cozinheiro", "name": "Cozinheiro(a)", "icon": "restaurant"},
    {"id": "marido_de_aluguel", "name": "Marido de Aluguel", "icon": "hammer"},
    {"id": "operador_maquinas", "name": "Operador de Máquinas", "icon": "cog"},
    {"id": "motorista_carreta", "name": "Motorista de Carreta", "icon": "bus"},
    {"id": "acougueiro", "name": "Açougueiro", "icon": "restaurant"},
    {"id": "seguranca", "name": "Segurança", "icon": "shield-checkmark"},
    {"id": "porteiro", "name": "Porteiro", "icon": "door"},
    {"id": "caseiro", "name": "Caseiro", "icon": "home"},
    {"id": "servente_obra", "name": "Servente de Obra", "icon": "construct"},
    {"id": "gesseiro", "name": "Gesseiro", "icon": "layers"},
    {"id": "soldador", "name": "Soldador", "icon": "flame"},
    {"id": "marceneiro", "name": "Marceneiro", "icon": "build"},
    {"id": "vidraceiro", "name": "Vidraceiro", "icon": "browsers"},
    {"id": "lavador_veiculos", "name": "Lavador de Veículos (Lava Jato)", "icon": "car-sport"},
    {"id": "funileiro", "name": "Funileiro", "icon": "car"},
    {"id": "borracheiro", "name": "Borracheiro", "icon": "ellipse"},
    {"id": "passadeira", "name": "Passadeira", "icon": "shirt"},
    {"id": "cuidador_idosos", "name": "Cuidador de Idosos", "icon": "heart"},
    {"id": "piscineiro", "name": "Piscineiro", "icon": "water"},
    {"id": "lavador_sofa", "name": "Lavador de Sofá/Estofado", "icon": "bed"},
    {"id": "tec_geladeira", "name": "Técnico de Geladeira/Freezer", "icon": "snow"},
    {"id": "tec_energia_solar", "name": "Técnico em Energia Solar", "icon": "sunny"},
    {"id": "bartender", "name": "Bartender", "icon": "wine"},
    {"id": "taxi_dog", "name": "Taxi-Dog", "icon": "paw"},
    {"id": "transporte_escolar", "name": "Transporte Escolar", "icon": "bus"},
]

# Bairros organizados por cidade
NEIGHBORHOODS_BY_CITY = {
    "tres_lagoas": [
        "Todos os bairros",
        "Alto da Boa Vista", "Bela Vista", "Carandá", "Centro", "Cinturão Verde",
        "Colinos", "Interlagos", "Jardim Alvorada", "Jardim Atenas", "Jardim Bela Vista",
        "Jardim Glória", "Jardim Guaporé", "Jardim Imperial", "Jardim Maristela",
        "Jardim Mirassol", "Jardim Morumbi", "Jardim Nova Americana", "Jardim Nova Ipanema",
        "Jardim Oiti", "Jardim Planalto", "Jardim Primavera", "Jardim Progresso",
        "Jardim Santa Aurélia", "Jardim Santa Júlia", "Jardim Vendrell", "Jardim Violetas",
        "JK", "Lapa", "Nossa Senhora Aparecida", "Nossa Senhora das Graças", "Nova Europa",
        "SetSul", "São Carlos", "São João", "São Jorge", "Vila Alegre", "Vila Cardoso",
        "Vila Carioca", "Vila Guanabara", "Vila Haro", "Vila Maria", "Vila Nova",
        "Vila Piloto", "Vila Popular", "Vila Santana", "Vila Verde", "Vila Viana"
    ],
    "andradina": [
        "Todos os bairros",
        "Centro", "Benfica", "Bom Jardim", "Catanduva", "Cidade Jardim",
        "Conjunto Habitacional Gasparelli", "Jardim Alvorada", "Jardim Brasil",
        "Jardim das Flores", "Jardim Europa", "Jardim Ipanema", "Jardim Noemia",
        "Jardim Olinda", "Jardim Palmeiras", "Jardim Paraíso", "Jardim Planalto",
        "Jardim Progresso", "Jardim Ribeiro", "Jardim Santa Cecília", "Jardim São Jorge",
        "Jardim São Paulo", "Jardim Stella Maris", "Parque das Nações", "Pereira Jordão",
        "Santo Antônio", "São Benedito", "São Joaquim", "Vila Alba", "Vila Mineira",
        "Vila Rica"
    ],
    "brasilandia": [
        "Todos os bairros",
        "Centro", "Bairro Alto", "Bela Vista", "Boa Esperança", "Conjunto Habitacional",
        "Jardim América", "Jardim Brasil", "Jardim das Flores", "Jardim Europa",
        "Jardim Planalto", "Jardim Primavera", "Nova Brasilândia", "Parque Industrial",
        "São Francisco", "Vila Nova", "Vila São Pedro"
    ],
    "selviria": [
        "Todos os bairros",
        "Centro", "Cohab", "Jardim Primavera", "Jardim São Paulo", "Vila Nova",
        "Vila Operária", "Vila Planalto"
    ],
    "agua_clara": [
        "Todos os bairros",
        "Centro", "Cohab", "Jardim América", "Jardim Bela Vista", "Jardim Brasil",
        "Jardim Planalto", "Jardim Primavera", "Nova Água Clara", "Vila Nova"
    ],
    "inocencia": [
        "Todos os bairros",
        "Centro", "Jardim América", "Jardim das Flores", "Jardim Europa",
        "Jardim Planalto", "Jardim Primavera", "Vila Nova", "Vila São Paulo"
    ],
    "paranaiba": [
        "Todos os bairros",
        "Alto da Colina", "Centro", "Cohab", "Jardim América", "Jardim Brasil",
        "Jardim das Flores", "Jardim Europa", "Jardim Guanabara", "Jardim Planalto",
        "Jardim Primavera", "Jardim Santa Mônica", "Jardim São Paulo", "Nova Paranaíba",
        "Vila Ipiranga", "Vila Nova", "Vila Operária"
    ],
    "aparecida_do_taboado": [
        "Todos os bairros",
        "Centro", "Cohab", "Jardim América", "Jardim Bela Vista", "Jardim Brasil",
        "Jardim das Flores", "Jardim Guanabara", "Jardim Planalto", "Jardim Primavera",
        "Jardim São Paulo", "Nova Aparecida", "Vila Nova", "Vila Operária"
    ],
    "ribas_do_rio_pardo": [
        "Todos os bairros",
        "Centro", "Cohab", "Jardim América", "Jardim Brasil", "Jardim das Flores",
        "Jardim Planalto", "Jardim Primavera", "Nova Ribas", "Vila Nova", "Vila Operária"
    ],
    "maringa": [
        "Todos os bairros",
        "Centro", "Zona 01", "Zona 02", "Zona 03", "Zona 04", "Zona 05", "Zona 06", "Zona 07",
        "Vila Morangueira", "Vila Santo Antônio", "Vila Nova", "Vila Esperança", "Vila Operária",
        "Jardim Alvorada", "Jardim América", "Jardim Aclimação", "Jardim Andrade", "Jardim Atami",
        "Jardim Brasília", "Jardim Canadá", "Jardim Dias", "Jardim Guaporé", "Jardim Imperial",
        "Jardim Indaiá", "Jardim Internorte", "Jardim Ipanema", "Jardim Itália", "Jardim Liberdade",
        "Jardim Mandacaru", "Jardim Maravilha", "Jardim Monte Carlo", "Jardim Novo Horizonte",
        "Jardim Olímpico", "Jardim Oásis", "Jardim Paris", "Jardim Paulista", "Jardim Pinheiros",
        "Jardim Real", "Jardim Santa Helena", "Jardim São Jorge", "Jardim São Silvestre",
        "Jardim Tóquio", "Jardim Tropical", "Jardim Universitário", "Parque Avenida",
        "Parque das Bandeiras", "Parque das Laranjeiras", "Parque Industrial", "Parque Residencial Cidade Nova",
        "Conjunto Habitacional Hermann Moraes Barros", "Conjunto Habitacional Iguatemi",
        "Conjunto Habitacional Sol Nascente", "Cidade Universitária", "Maringá Velho"
    ],
    "curitiba": [
        "Todos os bairros",
        "Centro", "Água Verde", "Alto Boqueirão", "Alto da Glória", "Alto da XV", "Ahú",
        "Abranches", "Atuba", "Augusta", "Bacacheri", "Bairro Alto", "Barreirinha", "Batel",
        "Bigorrilho", "Boa Vista", "Bom Retiro", "Boqueirão", "Butiatuvinha", "Cabral",
        "Cachoeira", "Cajuru", "Campina do Siqueira", "Campo Comprido", "Campo de Santana",
        "Capão Raso", "Capão da Imbuia", "Cascatinha", "Caximba", "Centro Cívico",
        "Cidade Industrial", "Cristo Rei", "Fanny", "Fazendinha", "Ganchinho", "Guabirotuba",
        "Guaíra", "Hauer", "Hugo Lange", "Jardim Botânico", "Jardim das Américas", "Jardim Social",
        "Juvevê", "Lamenha Pequena", "Lindóia", "Mercês", "Mossunguê", "Novo Mundo", "Orleans",
        "Parolin", "Pilarzinho", "Pinheirinho", "Portão", "Prado Velho", "Rebouças", "Riviera",
        "Santa Cândida", "Santa Felicidade", "Santa Quitéria", "Santo Inácio", "São Braz",
        "São Francisco", "São João", "São Lourenço", "São Miguel", "Seminário", "Sítio Cercado",
        "Taboão", "Tarumã", "Tatuquara", "Tingui", "Uberaba", "Umbará", "Vila Izabel",
        "Vista Alegre", "Xaxim", "Ecoville"
    ],
    "alto_taquari": [
        "Todos os bairros",
        "Centro", "Setor Comercial", "Setor Residencial Norte", "Setor Residencial Sul",
        "Vila Rural", "Área Industrial"
    ],
    "alto_araguaia": [
        "Todos os bairros",
        "Centro", "Buriti", "Barreiro", "Colônia do Ariranha", "Graciosa", "Paraíso",
        "Ribeirão Claro", "Setor Comercial", "Vila Nova", "Jardim América"
    ],
    "costa_rica": [
        "Todos os bairros",
        "Centro", "Baús", "Buritizal", "Jardim Eminassai", "Jardim São Francisco",
        "Vila Alvorada", "Vila Santana", "Flor do Campo", "Sonho Meu", "Jardim Planalto"
    ]
}

# Lista consolidada de todos os bairros (para compatibilidade)
NEIGHBORHOODS = NEIGHBORHOODS_BY_CITY["tres_lagoas"]

CITIES = [
    {"id": "tres_lagoas", "name": "Três Lagoas", "state": "MS"},
    {"id": "andradina", "name": "Andradina", "state": "SP"},
    {"id": "brasilandia", "name": "Brasilândia", "state": "MS"},
    {"id": "selviria", "name": "Selvíria", "state": "MS"},
    {"id": "agua_clara", "name": "Água Clara", "state": "MS"},
    {"id": "inocencia", "name": "Inocência", "state": "MS"},
    {"id": "paranaiba", "name": "Paranaíba", "state": "MS"},
    {"id": "aparecida_do_taboado", "name": "Aparecida do Taboado", "state": "MS"},
    {"id": "ribas_do_rio_pardo", "name": "Ribas do Rio Pardo", "state": "MS"},
    {"id": "maringa", "name": "Maringá", "state": "PR"},
    {"id": "curitiba", "name": "Curitiba", "state": "PR"},
    {"id": "alto_taquari", "name": "Alto Taquari", "state": "MT"},
    {"id": "alto_araguaia", "name": "Alto Araguaia", "state": "MT"},
    {"id": "costa_rica", "name": "Costa Rica", "state": "MS"},
]

# ======================== AUTH HELPERS ========================

async def get_session_token(request: Request) -> Optional[str]:
    """Get session token from cookie or Authorization header"""
    session_token = request.cookies.get("session_token")
    if session_token:
        return session_token
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        return auth_header.replace("Bearer ", "")
    return None

async def get_current_user(request: Request) -> Optional[User]:
    """Get current authenticated user"""
    session_token = await get_session_token(request)
    if not session_token:
        return None
    
    session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        return None
    
    expires_at = session["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        return None
    
    user_doc = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if user_doc:
        return User(**user_doc)
    return None

async def require_auth(request: Request) -> User:
    """Require authentication - raises 401 if not authenticated"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")
    return user

async def require_admin(request: Request) -> User:
    """Require admin authentication - raises 401/403 if not admin"""
    user = await require_auth(request)
    if user.email != ADMIN_EMAIL:
        logger.warning(f"Unauthorized admin access attempt by {user.email}")
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador")
    return user

# ======================== DATABASE INDEXES (on startup) ========================

@app.on_event("startup")
async def create_indexes():
    """Create MongoDB indexes for performance and data integrity"""
    try:
        # Users collection
        await db.users.create_index("user_id", unique=True)
        await db.users.create_index("email", unique=True)
        await db.users.create_index("push_token", sparse=True)
        
        # Providers collection
        await db.providers.create_index("user_id", unique=True)
        await db.providers.create_index("provider_id", unique=True)
        await db.providers.create_index("is_active")
        await db.providers.create_index("categories")
        await db.providers.create_index("city")
        await db.providers.create_index([("is_active", 1), ("blocked", 1)])
        
        # Sessions collection
        await db.user_sessions.create_index("session_token", unique=True)
        await db.user_sessions.create_index("user_id")
        await db.user_sessions.create_index("expires_at", expireAfterSeconds=0)
        
        # Reviews collection
        await db.reviews.create_index("provider_id")
        await db.reviews.create_index("user_id")
        
        # Reports collection
        await db.reports.create_index("status")
        await db.reports.create_index("provider_id")
        await db.reports.create_index("created_at")
        
        # WhatsApp contacts collection
        await db.whatsapp_contacts.create_index("user_id")
        await db.whatsapp_contacts.create_index("provider_id")
        await db.whatsapp_contacts.create_index([("user_id", 1), ("provider_id", 1)])
        
        # Notifications collection
        await db.notifications.create_index("user_id")
        await db.notifications.create_index([("user_id", 1), ("read", 1)])
        
        # Access logs - TTL index to auto-delete after 90 days
        await db.access_logs.create_index("timestamp", expireAfterSeconds=90*24*60*60)
        await db.access_logs.create_index([("date", 1), ("user_type", 1)])
        
        # User presence - TTL index to auto-delete stale entries after 1 day
        await db.user_presence.create_index("user_id", unique=True)
        await db.user_presence.create_index("last_seen", expireAfterSeconds=24*60*60)
        
        # App settings
        await db.app_settings.create_index("key", unique=True)
        
        # Banners
        await db.banners.create_index("is_active")
        
        logger.info("✅ MongoDB indexes created successfully")
    except Exception as e:
        logger.error(f"⚠️ Error creating indexes: {e}")

# ======================== BACKGROUND SCHEDULER ========================

async def scheduled_notifications_checker():
    """Background task that checks and sends scheduled notifications every 60 seconds"""
    logger.info("🔔 Scheduled notifications checker started")
    while True:
        try:
            await asyncio.sleep(60)  # Check every 60 seconds
            
            now = datetime.now(timezone(timedelta(hours=-4)))  # Brazil timezone (AMT/UTC-4 - Mato Grosso do Sul)
            current_hour = now.hour
            current_minute = now.minute
            today_str = now.strftime("%Y-%m-%d")
            
            # Find active notifications not yet sent today
            notifications = await db.scheduled_notifications.find({
                "is_active": True,
                "last_sent_date": {"$ne": today_str}
            }).to_list(100)
            
            for notif in notifications:
                try:
                    parts = notif["time"].split(":")
                    notif_hour = int(parts[0])
                    notif_minute = int(parts[1])
                    
                    # Check if current time matches (within 2-minute window)
                    notif_total_min = notif_hour * 60 + notif_minute
                    current_total_min = current_hour * 60 + current_minute
                    diff = current_total_min - notif_total_min
                    
                    # Only send if we're 0-2 minutes AFTER the scheduled time
                    if diff >= 0 and diff <= 2:
                        target = notif.get("target", "all")
                        push_tokens = set()
                        
                        if target in ["all", "providers"]:
                            providers = await db.providers.find(
                                {"push_token": {"$exists": True, "$ne": None, "$ne": ""}},
                                {"push_token": 1}
                            ).to_list(10000)
                            for p in providers:
                                if p.get("push_token"):
                                    push_tokens.add(p["push_token"])
                        
                        if target in ["all", "clients"]:
                            users = await db.users.find(
                                {"push_token": {"$exists": True, "$ne": None, "$ne": ""}},
                                {"push_token": 1}
                            ).to_list(10000)
                            for u in users:
                                if u.get("push_token"):
                                    push_tokens.add(u["push_token"])
                        
                        if push_tokens:
                            messages = []
                            for token in push_tokens:
                                if token.startswith("ExponentPushToken") or token.startswith("ExpoPushToken"):
                                    messages.append({
                                        "to": token,
                                        "sound": "default",
                                        "title": notif["title"],
                                        "body": notif["message"],
                                    })
                            
                            # Send in batches of 100
                            for i in range(0, len(messages), 100):
                                batch = messages[i:i+100]
                                try:
                                    async with httpx.AsyncClient() as client:
                                        await client.post(
                                            "https://exp.host/--/api/v2/push/send",
                                            json=batch,
                                            headers={
                                                "Accept": "application/json",
                                                "Content-Type": "application/json",
                                            },
                                            timeout=30
                                        )
                                except Exception as e:
                                    logger.error(f"Error sending push batch: {e}")
                            
                            # Mark as sent today
                            await db.scheduled_notifications.update_one(
                                {"notification_id": notif["notification_id"]},
                                {"$set": {"last_sent_date": today_str}}
                            )
                            logger.info(f"🔔 Scheduled notification sent: '{notif['title']}' at {notif['time']} to {len(push_tokens)} devices")
                        
                except Exception as e:
                    logger.error(f"Error processing scheduled notification: {e}")
                    
        except Exception as e:
            logger.error(f"Scheduler error: {e}")
            await asyncio.sleep(30)

@app.on_event("startup")
async def start_scheduler():
    """Start the background notification scheduler"""
    asyncio.create_task(scheduled_notifications_checker())


# ======================== AUTH ENDPOINTS ========================

@api_router.post("/auth/session")
@limiter.limit("10/minute")
async def exchange_session(request: Request, response: Response):
    """Exchange session_id for session_token"""
    session_id = request.headers.get("X-Session-ID")
    if not session_id:
        raise HTTPException(status_code=400, detail="Session ID não fornecido")
    
    async with httpx.AsyncClient() as client:
        try:
            auth_response = await client.get(
                f"{AUTH_BACKEND_URL}/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id}
            )
            if auth_response.status_code != 200:
                raise HTTPException(status_code=401, detail="Sessão inválida")
            
            user_data = auth_response.json()
        except Exception as e:
            logger.error(f"Auth error: {e}")
            raise HTTPException(status_code=500, detail="Erro na autenticação")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    existing_user = await db.users.find_one({"email": user_data["email"]}, {"_id": 0})
    
    if existing_user:
        user_id = existing_user["user_id"]
    else:
        new_user = User(
            user_id=user_id,
            email=user_data["email"],
            name=user_data["name"],
            picture=user_data.get("picture")
        )
        await db.users.insert_one(new_user.model_dump())
    
    session_token = user_data["session_token"]
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    session = UserSession(
        user_id=user_id,
        session_token=session_token,
        expires_at=expires_at
    )
    
    await db.user_sessions.delete_many({"user_id": user_id})
    await db.user_sessions.insert_one(session.model_dump())
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    
    return {"user": user_doc, "session_token": session_token}

@api_router.get("/auth/me")
async def get_me(request: Request):
    """Get current user info"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")
    
    provider = await db.providers.find_one({"user_id": user.user_id}, {"_id": 0})
    
    return {
        "user": user.model_dump(),
        "provider": provider
    }

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    """Logout user"""
    session_token = await get_session_token(request)
    if session_token:
        await db.user_sessions.delete_many({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Logout realizado com sucesso"}

# ======================== CATEGORIES & NEIGHBORHOODS ========================

@api_router.get("/categories")
async def get_categories():
    """Get all service categories"""
    return CATEGORIES

@api_router.get("/neighborhoods")
async def get_neighborhoods(city: Optional[str] = None):
    """Get neighborhoods by city. If no city specified, returns all neighborhoods."""
    if city and city in NEIGHBORHOODS_BY_CITY:
        return NEIGHBORHOODS_BY_CITY[city]
    # Se não especificou cidade, retorna todos os bairros de todas as cidades
    return NEIGHBORHOODS

@api_router.get("/cities")
async def get_cities():
    """Get all available cities"""
    return CITIES

# ======================== SUBSCRIPTION EXPIRATION ========================

async def check_and_expire_subscriptions():
    """Check for expired subscriptions and deactivate them"""
    now = datetime.now(timezone.utc)
    
    # Find providers with expired subscriptions
    expired_providers = await db.providers.find({
        "subscription_status": "active",
        "subscription_expires_at": {"$lt": now}
    }).to_list(1000)
    
    expired_count = 0
    for provider in expired_providers:
        # Update provider status to expired
        await db.providers.update_one(
            {"provider_id": provider["provider_id"]},
            {"$set": {
                "subscription_status": "expired",
                "is_active": False
            }}
        )
        
        # Update subscription record
        await db.subscriptions.update_one(
            {"provider_id": provider["provider_id"], "status": "active"},
            {"$set": {"status": "expired"}}
        )
        
        expired_count += 1
        logger.info(f"Subscription expired for provider: {provider['provider_id']} - {provider['name']}")
    
    return expired_count

# ======================== PROVIDERS ========================

@api_router.get("/providers")
async def get_providers(
    category: Optional[str] = None,
    neighborhood: Optional[str] = None,
    city: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 20
):
    """Get all active providers with optional filters and pagination"""
    # First, check and expire any overdue subscriptions
    await check_and_expire_subscriptions()
    
    # Sanitize pagination params
    page = max(1, page)
    limit = min(max(1, limit), 100)  # Max 100 per page
    skip = (page - 1) * limit
    
    query = {"is_active": True, "subscription_status": "active"}
    
    if category:
        # Search in categories array
        query["categories"] = category
    if city:
        # Search in cities array
        query["cities"] = city
    if neighborhood and neighborhood != "Todos os bairros":
        # Include providers that serve this specific neighborhood OR all neighborhoods
        query["neighborhood"] = {"$in": [neighborhood, "Todos os bairros"]}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    total = await db.providers.count_documents(query)
    providers = await db.providers.find(query, {"_id": 0}).sort("average_rating", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "providers": providers,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit  # ceil division
    }

@api_router.get("/providers/{provider_id}")
async def get_provider(provider_id: str):
    """Get a specific provider by ID"""
    provider = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    return provider

@api_router.post("/providers", response_model=Provider)
@limiter.limit("3/minute")
async def create_provider(provider_data: ProviderCreate, request: Request):
    """Create a new provider profile (requires auth)"""
    user = await require_auth(request)
    
    existing = await db.providers.find_one({"user_id": user.user_id})
    if existing:
        raise HTTPException(status_code=400, detail="Você já possui um perfil de prestador")
    
    # Upload images to Cloudinary
    provider_dict = provider_data.model_dump()
    if provider_dict.get('profile_image'):
        provider_dict['profile_image'] = await upload_to_cloudinary(
            provider_dict['profile_image'], 
            folder="achaservico/profiles"
        )
    if provider_dict.get('service_photos'):
        provider_dict['service_photos'] = await upload_images_to_cloudinary(
            provider_dict['service_photos'],
            folder="achaservico/services"
        )
    
    # ESTRATÉGIA 100% GRÁTIS: Novos prestadores começam ativos sem data de expiração
    provider = Provider(
        user_id=user.user_id,
        subscription_status="active",  # Já começa ativo
        subscription_expires_at=None,  # Sem expiração - gratuito permanente
        is_active=True,
        **provider_dict
    )
    
    await db.providers.insert_one(provider.model_dump())
    
    # Criar registro de assinatura (sem expiração)
    subscription = Subscription(
        provider_id=provider.provider_id,
        user_id=user.user_id,
        status="active",
        payment_method="free",  # Gratuito
        started_at=datetime.now(timezone.utc),
        expires_at=None  # Sem expiração
    )
    await db.subscriptions.insert_one(subscription.model_dump())
    
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"is_provider": True}}
    )
    
    # Send email notification to admin about new provider registration (non-blocking)
    try:
        await send_admin_notification_email(
            provider_name=provider.name,
            provider_email=user.email,
            categories=provider.categories,
            cities=provider.cities,
            phone=provider.phone
        )
    except Exception as e:
        # Log error but don't fail the registration
        print(f"Warning: Failed to send admin notification email: {e}")
    
    return provider

@api_router.put("/providers/{provider_id}")
async def update_provider(provider_id: str, provider_data: ProviderUpdate, request: Request):
    """Update provider profile (requires auth and ownership)"""
    user = await require_auth(request)
    
    provider = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    if provider["user_id"] != user.user_id:
        raise HTTPException(status_code=403, detail="Você não tem permissão para editar este perfil")
    
    update_data = {k: v for k, v in provider_data.model_dump().items() if v is not None}
    
    # Upload new images to Cloudinary
    if update_data.get('profile_image') and update_data['profile_image'].startswith('data:'):
        update_data['profile_image'] = await upload_to_cloudinary(
            update_data['profile_image'],
            folder="achaservico/profiles"
        )
    if update_data.get('service_photos'):
        # Only upload new base64 images, keep existing URLs
        new_photos = []
        for photo in update_data['service_photos']:
            if photo.startswith('data:') or photo.startswith('data:image'):
                uploaded = await upload_to_cloudinary(photo, folder="achaservico/services")
                new_photos.append(uploaded)
            else:
                new_photos.append(photo)
        update_data['service_photos'] = new_photos
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": update_data}
    )
    
    updated = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0})
    return updated

@api_router.post("/providers/{provider_id}/toggle-availability")
async def toggle_provider_availability(provider_id: str, request: Request):
    """Toggle provider's visibility on the platform"""
    user = await require_auth(request)
    
    provider = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    if provider["user_id"] != user.user_id:
        raise HTTPException(status_code=403, detail="Você não tem permissão para alterar este perfil")
    
    # Toggle visibility (is_active controls if provider appears in search)
    # Blocked providers cannot toggle visibility
    if provider.get("blocked"):
        raise HTTPException(status_code=403, detail="Seu perfil está bloqueado. Entre em contato com o suporte.")
    
    new_status = not provider.get("is_active", True)
    
    await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": {"is_active": new_status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"success": True, "is_active": new_status}

# ======================== FAVORITES ========================

@api_router.get("/users/favorites")
async def get_user_favorites(request: Request):
    """Get user's favorite providers"""
    user = await require_auth(request)
    
    user_data = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    favorite_ids = user_data.get("favorite_providers", []) if user_data else []
    
    if not favorite_ids:
        return []
    
    # Get provider details for favorites
    favorites = await db.providers.find(
        {"provider_id": {"$in": favorite_ids}, "is_active": True},
        {"_id": 0, "service_photos": 0}
    ).to_list(100)
    
    return favorites

@api_router.post("/users/favorites/{provider_id}")
async def toggle_favorite(provider_id: str, request: Request):
    """Add or remove a provider from favorites"""
    user = await require_auth(request)
    
    # Check if provider exists
    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    # Get current favorites
    user_data = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    current_favorites = user_data.get("favorite_providers", []) if user_data else []
    
    # Toggle favorite
    if provider_id in current_favorites:
        current_favorites.remove(provider_id)
        is_favorite = False
    else:
        current_favorites.append(provider_id)
        is_favorite = True
        
        # Save notification to database (appears in notification center / bell icon)
        notification_entry = Notification(
            user_id=provider.get("user_id"),
            title="Novo favorito! ⭐",
            message="Alguém adicionou você aos favoritos!",
            notification_type="favorite"
        )
        await db.notifications.insert_one(notification_entry.model_dump())
        logger.info(f"Saved favorite notification for provider {provider_id}")
        
        # Send push notification to provider when favorited
        provider_user = await db.users.find_one({"user_id": provider.get("user_id")})
        if provider_user and provider_user.get("push_token"):
            try:
                await send_push_notification(
                    push_token=provider_user["push_token"],
                    title="Novo favorito! ⭐",
                    body=f"Alguém adicionou você aos favoritos!",
                    data={"type": "new_favorite", "notification_id": notification_entry.notification_id}
                )
                logger.info(f"Sent favorite push notification to provider {provider_id}")
            except Exception as e:
                logger.error(f"Error sending favorite notification: {e}")
    
    # Update user
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"favorite_providers": current_favorites}}
    )
    
    return {"success": True, "is_favorite": is_favorite}

@api_router.get("/users/favorites/check/{provider_id}")
async def check_favorite(provider_id: str, request: Request):
    """Check if a provider is in user's favorites"""
    user = await require_auth(request)
    
    user_data = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    favorites = user_data.get("favorite_providers", []) if user_data else []
    
    return {"is_favorite": provider_id in favorites}

# ======================== REVIEWS ========================

@api_router.get("/providers/{provider_id}/reviews")
async def get_provider_reviews(provider_id: str):
    """Get all reviews for a provider (ANONYMOUS - no user identification)"""
    reviews = await db.reviews.find({"provider_id": provider_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Remove user identification for anonymity
    anonymous_reviews = []
    for review in reviews:
        anonymous_reviews.append({
            "review_id": review.get("review_id"),
            "provider_id": review.get("provider_id"),
            "rating": review.get("rating"),
            "comment": review.get("comment"),
            "is_verified": review.get("is_verified", True),
            "created_at": review.get("created_at")
            # user_id and user_name are NOT included
        })
    
    return anonymous_reviews

@api_router.post("/providers/{provider_id}/contact")
async def register_whatsapp_contact(provider_id: str, request: Request):
    """Register when a user clicks on WhatsApp button (enables reviews)"""
    user = await get_current_user(request)
    
    if not user:
        # Allow anonymous contact but they won't be able to review
        return {"success": True, "can_review": False, "message": "Contato registrado (faça login para avaliar depois)"}
    
    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    # Check if contact already exists
    existing_contact = await db.whatsapp_contacts.find_one({
        "user_id": user.user_id,
        "provider_id": provider_id
    })
    
    if not existing_contact:
        contact = WhatsAppContact(
            user_id=user.user_id,
            provider_id=provider_id
        )
        await db.whatsapp_contacts.insert_one(contact.model_dump())
    
    return {"success": True, "can_review": True, "message": "Contato registrado, você poderá avaliar este prestador"}

@api_router.get("/providers/{provider_id}/can-review")
async def check_can_review(provider_id: str, request: Request):
    """Check if user can review this provider (must have contacted via WhatsApp and not be the provider)"""
    user = await get_current_user(request)
    
    if not user:
        return {"can_review": False, "reason": "not_authenticated"}
    
    # Check if user is the provider (block self-review)
    provider = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0, "user_id": 1})
    if provider and provider.get("user_id") == user.user_id:
        return {"can_review": False, "reason": "self_review_not_allowed"}
    
    # Check if user already reviewed
    existing_review = await db.reviews.find_one({
        "provider_id": provider_id,
        "user_id": user.user_id
    })
    if existing_review:
        return {"can_review": False, "reason": "already_reviewed"}
    
    # Check if user has contacted this provider
    contact = await db.whatsapp_contacts.find_one({
        "user_id": user.user_id,
        "provider_id": provider_id
    })
    
    if not contact:
        return {"can_review": False, "reason": "no_contact"}
    
    return {"can_review": True, "reason": "eligible"}

@api_router.get("/users/contact-history")
async def get_contact_history(request: Request):
    """Get list of providers the user has contacted (history)"""
    user = await require_auth(request)
    
    # Get all contacts for this user (sorted by most recent first)
    contacts = await db.whatsapp_contacts.find(
        {"user_id": user.user_id}
    ).to_list(100)
    
    # Sort by contacted_at or created_at (for old records)
    contacts.sort(key=lambda x: x.get("contacted_at") or x.get("created_at") or datetime.min, reverse=True)
    
    if not contacts:
        return []
    
    # Get provider details for each contact
    provider_ids = [c["provider_id"] for c in contacts]
    providers = await db.providers.find(
        {"provider_id": {"$in": provider_ids}},
        {"_id": 0, "profile_image": 1, "provider_id": 1, "name": 1, "categories": 1, "phone": 1, "neighborhood": 1}
    ).to_list(100)
    
    # Create a map for quick lookup
    provider_map = {p["provider_id"]: p for p in providers}
    
    # Build result with contact date
    result = []
    for contact in contacts:
        provider = provider_map.get(contact["provider_id"])
        if provider:
            # Try contacted_at first, fall back to created_at for old records
            contacted_at = contact.get("contacted_at") or contact.get("created_at")
            # Convert datetime to ISO string if needed
            if contacted_at and hasattr(contacted_at, 'isoformat'):
                contacted_at = contacted_at.isoformat()
            result.append({
                **provider,
                "contacted_at": contacted_at
            })
    
    return result

@api_router.post("/reviews")
async def create_review(review_data: ReviewCreate, request: Request):
    """Create a review for a provider (requires auth + WhatsApp contact)"""
    user = await require_auth(request)
    
    provider = await db.providers.find_one({"provider_id": review_data.provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    # Check if user has contacted this provider via WhatsApp
    contact = await db.whatsapp_contacts.find_one({
        "user_id": user.user_id,
        "provider_id": review_data.provider_id
    })
    if not contact:
        raise HTTPException(
            status_code=403, 
            detail="Você precisa entrar em contato com o prestador pelo WhatsApp antes de avaliar"
        )
    
    # Check if already reviewed
    existing_review = await db.reviews.find_one({
        "provider_id": review_data.provider_id,
        "user_id": user.user_id
    })
    if existing_review:
        raise HTTPException(status_code=400, detail="Você já avaliou este prestador")
    
    if review_data.rating < 1 or review_data.rating > 5:
        raise HTTPException(status_code=400, detail="Avaliação deve ser entre 1 e 5")
    
    review = Review(
        provider_id=review_data.provider_id,
        user_id=user.user_id,
        user_name=user.name,  # Stored but not shown to provider
        rating=review_data.rating,
        comment=review_data.comment,
        is_verified=True  # Verified because they contacted via WhatsApp
    )
    
    await db.reviews.insert_one(review.model_dump())
    
    # Update provider average rating using aggregation (optimized)
    pipeline = [
        {"$match": {"provider_id": review_data.provider_id}},
        {"$group": {
            "_id": None,
            "avg_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1}
        }}
    ]
    result = await db.reviews.aggregate(pipeline).to_list(1)
    
    if result:
        avg_rating = result[0]["avg_rating"]
        total_reviews = result[0]["total_reviews"]
    else:
        avg_rating = review_data.rating
        total_reviews = 1
    
    # Verificar automaticamente: 5+ avaliações positivas (4+ estrelas)
    positive_reviews_pipeline = [
        {"$match": {"provider_id": review_data.provider_id, "rating": {"$gte": 4}}},
        {"$count": "count"}
    ]
    positive_result = await db.reviews.aggregate(positive_reviews_pipeline).to_list(1)
    positive_count = positive_result[0]["count"] if positive_result else 0
    is_verified = positive_count >= 5
    
    await db.providers.update_one(
        {"provider_id": review_data.provider_id},
        {"$set": {
            "average_rating": round(avg_rating, 1),
            "total_reviews": total_reviews,
            "is_verified": is_verified
        }}
    )
    
    # Return anonymous version (without user_id and user_name)
    return {
        "review_id": review.review_id,
        "provider_id": review.provider_id,
        "rating": review.rating,
        "comment": review.comment,
        "is_verified": review.is_verified,
        "created_at": review.created_at.isoformat()
    }

# ======================== PIX MANUAL SUBSCRIPTIONS ========================

@api_router.get("/payments/pix-info")
async def get_pix_info():
    """Get PIX information for manual payment"""
    return {
        "pix_key": PIX_KEY,
        "pix_key_type": PIX_KEY_TYPE,
        "pix_key_formatted": "499.586.888-75",
        "receiver_name": PIX_RECEIVER_NAME,
        "amount": 9.99,
        "description": "Assinatura Mensal AchaServico"
    }

@api_router.post("/subscriptions/create")
async def create_subscription(request: Request):
    """Create a pending subscription for PIX payment"""
    user = await require_auth(request)
    
    provider = await db.providers.find_one({"user_id": user.user_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=400, detail="Você precisa criar um perfil de prestador primeiro")
    
    # Check if there's already a pending subscription
    existing = await db.subscriptions.find_one({
        "provider_id": provider["provider_id"],
        "status": {"$in": ["pending", "active"]}
    })
    
    if existing and existing.get("status") == "active":
        raise HTTPException(status_code=400, detail="Você já possui uma assinatura ativa")
    
    # Create or update pending subscription
    subscription = Subscription(
        provider_id=provider["provider_id"],
        user_id=user.user_id,
        status="pending",
        payment_method="pix_manual"
    )
    
    if existing:
        await db.subscriptions.update_one(
            {"provider_id": provider["provider_id"]},
            {"$set": {"status": "pending", "created_at": datetime.now(timezone.utc)}}
        )
    else:
        await db.subscriptions.insert_one(subscription.model_dump())
    
    return {
        "success": True,
        "subscription_id": subscription.subscription_id,
        "pix_key": PIX_KEY,
        "pix_key_formatted": "(66) 99684-1531",
        "amount": 9.99,
        "message": "Faça o PIX e aguarde a confirmação"
    }

@api_router.get("/subscriptions/status")
async def get_subscription_status(request: Request):
    """Get current subscription status"""
    user = await require_auth(request)
    
    provider = await db.providers.find_one({"user_id": user.user_id}, {"_id": 0})
    if not provider:
        return {"has_provider": False, "subscription": None}
    
    subscription = await db.subscriptions.find_one(
        {"provider_id": provider["provider_id"]},
        {"_id": 0}
    )
    
    return {
        "has_provider": True,
        "provider": provider,
        "subscription": subscription
    }

@api_router.post("/subscriptions/activate")
async def activate_subscription_manual(request: Request):
    """Manually activate subscription (for testing or manual approval)"""
    user = await require_auth(request)
    
    provider = await db.providers.find_one({"user_id": user.user_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=400, detail="Prestador não encontrado")
    
    expires_at = datetime.now(timezone.utc) + timedelta(days=30)
    
    # Update subscription
    await db.subscriptions.update_one(
        {"provider_id": provider["provider_id"], "status": "pending"},
        {"$set": {
            "status": "active",
            "started_at": datetime.now(timezone.utc),
            "expires_at": expires_at
        }}
    )
    
    # Activate provider
    await db.providers.update_one(
        {"provider_id": provider["provider_id"]},
        {"$set": {
            "subscription_status": "active",
            "subscription_expires_at": expires_at,
            "is_active": True
        }}
    )
    
    return {"success": True, "message": "Assinatura ativada com sucesso!", "expires_at": expires_at.isoformat()}


# ======================== SCHEDULED NOTIFICATIONS ========================

class ScheduledNotificationCreate(BaseModel):
    time: str  # Format "HH:MM"
    title: str
    message: str
    target: str = "all"  # all, providers, clients

@api_router.get("/admin/scheduled-notifications")
async def get_scheduled_notifications(request: Request):
    """Get all scheduled notifications"""
    await require_admin(request)
    notifications = await db.scheduled_notifications.find({}, {"_id": 0}).to_list(100)
    return notifications

@api_router.post("/admin/scheduled-notifications")
async def create_scheduled_notification(request: Request, data: ScheduledNotificationCreate):
    """Create a new scheduled notification"""
    await require_admin(request)
    
    # Validate time format
    try:
        parts = data.time.split(":")
        hour = int(parts[0])
        minute = int(parts[1])
        if hour < 0 or hour > 23 or minute < 0 or minute > 59:
            raise ValueError()
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Formato de hora inválido. Use HH:MM (ex: 10:00)")
    
    notification = {
        "notification_id": str(uuid.uuid4()),
        "time": data.time,
        "title": data.title,
        "message": data.message,
        "target": data.target,
        "is_active": True,
        "last_sent_date": None,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.scheduled_notifications.insert_one(notification)
    del notification["_id"]
    return notification

@api_router.put("/admin/scheduled-notifications/{notification_id}")
async def update_scheduled_notification(request: Request, notification_id: str):
    """Update a scheduled notification (toggle active, edit fields)"""
    await require_admin(request)
    body = await request.json()
    
    update_fields = {}
    if "is_active" in body:
        update_fields["is_active"] = body["is_active"]
    if "time" in body:
        update_fields["time"] = body["time"]
    if "title" in body:
        update_fields["title"] = body["title"]
    if "message" in body:
        update_fields["message"] = body["message"]
    if "target" in body:
        update_fields["target"] = body["target"]
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    
    update_fields["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.scheduled_notifications.update_one(
        {"notification_id": notification_id},
        {"$set": update_fields}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notificação não encontrada")
    
    return {"success": True, "message": "Notificação atualizada"}

@api_router.delete("/admin/scheduled-notifications/{notification_id}")
async def delete_scheduled_notification(request: Request, notification_id: str):
    """Delete a scheduled notification"""
    await require_admin(request)
    
    result = await db.scheduled_notifications.delete_one({"notification_id": notification_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notificação não encontrada")
    
    return {"success": True, "message": "Notificação excluída"}

@api_router.post("/admin/scheduled-notifications/{notification_id}/send-now")
async def send_notification_now(request: Request, notification_id: str):
    """Force send a scheduled notification immediately"""
    await require_admin(request)
    
    notif = await db.scheduled_notifications.find_one({"notification_id": notification_id})
    if not notif:
        raise HTTPException(status_code=404, detail="Notificação não encontrada")
    
    today_str = datetime.now(timezone(timedelta(hours=-4))).strftime("%Y-%m-%d")
    await send_scheduled_push(notif, today_str)
    
    return {"success": True, "message": f"Notificação '{notif['title']}' enviada!"}

@api_router.post("/cron/send-scheduled-notifications")
@api_router.get("/cron/send-scheduled-notifications")
async def send_scheduled_notifications(background_tasks: BackgroundTasks):
    """Cron endpoint - responds immediately, processes notifications in background.
    Supports both GET and POST for cron service compatibility."""
    now = datetime.now(timezone(timedelta(hours=-4)))  # Brazil timezone (AMT/UTC-4 - Mato Grosso do Sul)
    current_time = now.strftime("%H:%M")
    logger.info(f"🔔 Cron ping received at {current_time}")
    
    # Process notifications in background to avoid timeout
    background_tasks.add_task(_process_scheduled_notifications)
    
    return {"success": True, "message": "Processing notifications in background", "checked_at": current_time}

async def _process_scheduled_notifications():
    """Background task to check and send scheduled notifications."""
    try:
        now = datetime.now(timezone(timedelta(hours=-4)))  # Brazil timezone (AMT/UTC-4 - Mato Grosso do Sul)
        current_hour = now.hour
        current_minute = now.minute
        today_str = now.strftime("%Y-%m-%d")
        
        # Find active notifications that match current time (within 5-minute window)
        notifications = await db.scheduled_notifications.find({
            "is_active": True,
            "last_sent_date": {"$ne": today_str}
        }).to_list(100)
        
        sent_count = 0
        for notif in notifications:
            try:
                parts = notif["time"].split(":")
                notif_hour = int(parts[0])
                notif_minute = int(parts[1])
                
                # Check if within 5-minute window
                notif_total_min = notif_hour * 60 + notif_minute
                current_total_min = current_hour * 60 + current_minute
                diff = abs(current_total_min - notif_total_min)
                
                if diff <= 5 or diff >= (24*60 - 5):  # Handle midnight wrap
                    # Determine target tokens
                    target = notif.get("target", "all")
                    push_tokens = set()
                    
                    if target in ["all", "providers"]:
                        providers = await db.providers.find(
                            {"push_token": {"$exists": True, "$ne": None, "$ne": ""}},
                            {"push_token": 1}
                        ).to_list(10000)
                        for p in providers:
                            if p.get("push_token"):
                                push_tokens.add(p["push_token"])
                    
                    if target in ["all", "clients"]:
                        users = await db.users.find(
                            {"push_token": {"$exists": True, "$ne": None, "$ne": ""}},
                            {"push_token": 1}
                        ).to_list(10000)
                        for u in users:
                            if u.get("push_token"):
                                push_tokens.add(u["push_token"])
                    
                    if push_tokens:
                        # Send via Expo Push
                        messages = []
                        for token in push_tokens:
                            if token.startswith("ExponentPushToken") or token.startswith("ExpoPushToken"):
                                messages.append({
                                    "to": token,
                                    "sound": "default",
                                    "title": notif["title"],
                                    "body": notif["message"],
                                })
                        
                        # Send in batches of 100
                        for i in range(0, len(messages), 100):
                            batch = messages[i:i+100]
                            try:
                                async with httpx.AsyncClient() as client:
                                    await client.post(
                                        "https://exp.host/--/api/v2/push/send",
                                        json=batch,
                                        headers={
                                            "Accept": "application/json",
                                            "Content-Type": "application/json",
                                        },
                                        timeout=30
                                    )
                            except Exception as e:
                                logger.error(f"Error sending push batch: {e}")
                        
                        # Mark as sent today
                        await db.scheduled_notifications.update_one(
                            {"notification_id": notif["notification_id"]},
                            {"$set": {"last_sent_date": today_str}}
                        )
                        sent_count += 1
                        logger.info(f"✅ Scheduled notification sent: '{notif['title']}' to {len(push_tokens)} devices")
                        
            except Exception as e:
                logger.error(f"Error processing scheduled notification: {e}")
        
        if sent_count > 0:
            logger.info(f"🔔 Cron job completed: {sent_count} notifications sent")
        
    except Exception as e:
        logger.error(f"Error in background notification processing: {e}")


# ======================== MAINTENANCE MODE ========================

@api_router.get("/maintenance/status")
async def get_maintenance_status(emergency_off: str = None):
    """Public endpoint - check if app is in maintenance mode.
    Admin can pass ?emergency_off=achaservico2026 to disable maintenance."""
    # Emergency maintenance disable for admin lockout situations
    if emergency_off == "achaservico2026":
        await db.app_settings.update_one(
            {"key": "maintenance"},
            {"$set": {"active": False, "updated_at": datetime.now(timezone.utc)}},
            upsert=True
        )
        logger.info("⚠️ Maintenance mode disabled via emergency parameter")
        return {"active": False, "message": "Manutenção desativada via emergência"}
    
    # Piggyback: trigger scheduled notifications check (non-blocking)
    try:
        asyncio.create_task(check_and_send_scheduled_notifications())
    except Exception:
        pass
    
    maintenance = await db.app_settings.find_one({"key": "maintenance"})
    if maintenance and maintenance.get("active"):
        return {
            "active": True,
            "message": maintenance.get("message", "Estamos em manutenção. Voltamos em breve!")
        }
    return {"active": False, "message": ""}

@api_router.post("/admin/maintenance/toggle")
@limiter.limit("10/minute")
async def toggle_maintenance(request: Request):
    """Admin: Toggle maintenance mode on/off"""
    await require_admin(request)
    body = await request.json()
    active = body.get("active", False)
    message = body.get("message", "Estamos realizando uma manutenção programada. Voltamos em breve!")
    
    await db.app_settings.update_one(
        {"key": "maintenance"},
        {"$set": {
            "key": "maintenance",
            "active": active,
            "message": message,
            "updated_at": datetime.now(timezone.utc)
        }},
        upsert=True
    )
    
    status_text = "ativada" if active else "desativada"
    logger.info(f"Maintenance mode {status_text}")
    return {"success": True, "active": active, "message": f"Manutenção {status_text}"}

# ======================== HEARTBEAT & ONLINE TRACKING ========================

ONLINE_THRESHOLD_MINUTES = 5  # User is "online" if heartbeat within last 5 minutes

@api_router.post("/heartbeat")
@limiter.limit("35/minute")
async def user_heartbeat(request: Request):
    """Track user presence and log access"""
    user = await get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Não autenticado")
    
    now = datetime.now(timezone.utc)
    today_str = now.strftime("%Y-%m-%d")
    
    # Determine if user is a provider or client
    provider = await db.providers.find_one({"user_id": user.user_id, "is_active": True})
    user_type = "provider" if provider else "client"
    
    # 1. Log this access (every call counts as 1 access)
    await db.access_logs.insert_one({
        "user_id": user.user_id,
        "user_type": user_type,
        "date": today_str,
        "timestamp": now
    })
    
    # 2. Update presence (upsert - marks user as online)
    await db.user_presence.update_one(
        {"user_id": user.user_id},
        {"$set": {
            "user_id": user.user_id,
            "user_type": user_type,
            "user_name": user.name,
            "last_seen": now
        }},
        upsert=True
    )
    
    # 3. Piggyback: check scheduled notifications (non-blocking)
    try:
        asyncio.create_task(check_and_send_scheduled_notifications())
    except Exception:
        pass
    
    return {"status": "ok"}

async def check_and_send_scheduled_notifications():
    """Check and send any pending scheduled notifications"""
    try:
        now = datetime.now(timezone(timedelta(hours=-4)))  # Brazil timezone (AMT/UTC-4 - Mato Grosso do Sul)
        current_hour = now.hour
        current_minute = now.minute
        today_str = now.strftime("%Y-%m-%d")
        
        # Rate-limit: only check once per 5-minute window
        lock_key = f"notif_lock_{today_str}_{current_hour}_{current_minute // 5}"
        lock = await db.app_settings.find_one({"key": lock_key})
        if lock:
            return  # Already checked this window
        await db.app_settings.update_one(
            {"key": lock_key},
            {"$set": {"key": lock_key, "ts": now}},
            upsert=True
        )
        
        notifications = await db.scheduled_notifications.find({
            "is_active": True,
            "last_sent_date": {"$ne": today_str}
        }).to_list(100)
        
        for notif in notifications:
            try:
                parts = notif["time"].split(":")
                notif_hour = int(parts[0])
                notif_minute = int(parts[1])
                
                notif_total_min = notif_hour * 60 + notif_minute
                current_total_min = current_hour * 60 + current_minute
                diff = current_total_min - notif_total_min
                
                # Send if we're 0-10 minutes AFTER the scheduled time
                if diff >= 0 and diff <= 10:
                    await send_scheduled_push(notif, today_str)
                    logger.info(f"🔔 Auto-sent: '{notif['title']}' at {notif['time']}")
            except Exception as e:
                logger.error(f"Error checking notification: {e}")
    except Exception as e:
        logger.error(f"Scheduled check error: {e}")

async def send_scheduled_push(notif: dict, today_str: str):
    """Send a scheduled push notification to target users"""
    target = notif.get("target", "all")
    push_tokens = set()
    
    if target in ["all", "providers"]:
        providers = await db.providers.find(
            {"push_token": {"$exists": True, "$ne": None, "$ne": ""}},
            {"push_token": 1}
        ).to_list(10000)
        for p in providers:
            if p.get("push_token"):
                push_tokens.add(p["push_token"])
    
    if target in ["all", "clients"]:
        users = await db.users.find(
            {"push_token": {"$exists": True, "$ne": None, "$ne": ""}},
            {"push_token": 1}
        ).to_list(10000)
        for u in users:
            if u.get("push_token"):
                push_tokens.add(u["push_token"])
    
    if push_tokens:
        messages = []
        for token in push_tokens:
            if token.startswith("ExponentPushToken") or token.startswith("ExpoPushToken"):
                messages.append({
                    "to": token,
                    "sound": "default",
                    "title": notif["title"],
                    "body": notif["message"],
                })
        
        for i in range(0, len(messages), 100):
            batch = messages[i:i+100]
            try:
                async with httpx.AsyncClient() as client:
                    await client.post(
                        "https://exp.host/--/api/v2/push/send",
                        json=batch,
                        headers={
                            "Accept": "application/json",
                            "Content-Type": "application/json",
                        },
                        timeout=30
                    )
            except Exception as e:
                logger.error(f"Error sending push batch: {e}")
        
        await db.scheduled_notifications.update_one(
            {"notification_id": notif["notification_id"]},
            {"$set": {"last_sent_date": today_str}}
        )
        logger.info(f"🔔 Sent: '{notif['title']}' at {notif['time']} to {len(push_tokens)} devices")

@api_router.get("/admin/online-stats")
async def get_online_stats(request: Request):
    """Get real-time online users and daily access counts"""
    await require_admin(request)
    now = datetime.now(timezone.utc)
    today_str = now.strftime("%Y-%m-%d")
    threshold = now - timedelta(minutes=ONLINE_THRESHOLD_MINUTES)
    
    # --- Online NOW (heartbeat within last 5 minutes) ---
    online_providers = await db.user_presence.count_documents({
        "user_type": "provider",
        "last_seen": {"$gte": threshold}
    })
    online_clients = await db.user_presence.count_documents({
        "user_type": "client",
        "last_seen": {"$gte": threshold}
    })
    online_total = online_providers + online_clients
    
    # --- Daily accesses (total visits today, including duplicates) ---
    daily_total = await db.access_logs.count_documents({"date": today_str})
    daily_providers = await db.access_logs.count_documents({"date": today_str, "user_type": "provider"})
    daily_clients = await db.access_logs.count_documents({"date": today_str, "user_type": "client"})
    
    return {
        "online_now": {
            "total": online_total,
            "providers": online_providers,
            "clients": online_clients
        },
        "daily_accesses": {
            "total": daily_total,
            "providers": daily_providers,
            "clients": daily_clients
        }
    }

# ======================== ADMIN ENDPOINTS ========================

@api_router.get("/admin/stats")
async def get_admin_stats(request: Request):
    """Get comprehensive dashboard statistics"""
    await require_admin(request)
    # Check for expired subscriptions first
    await check_and_expire_subscriptions()
    
    total_users = await db.users.count_documents({})
    total_providers = await db.providers.count_documents({})
    active_subscriptions = await db.providers.count_documents({"subscription_status": "active"})
    
    # Count pending as: providers with inactive subscription + subscriptions marked as pending
    inactive_providers = await db.providers.count_documents({"subscription_status": "inactive"})
    pending_subs = await db.subscriptions.count_documents({"status": "pending"})
    pending_subscriptions = inactive_providers + pending_subs
    
    expired_subscriptions = await db.providers.count_documents({"subscription_status": "expired"})
    total_reviews = await db.reviews.count_documents({})
    pending_reports = await db.reports.count_documents({"status": "pending"})
    total_reports = await db.reports.count_documents({})
    
    # Advanced stats
    total_contacts = await db.whatsapp_contacts.count_documents({})
    total_favorites = 0
    users_with_favs = await db.users.find({"favorites": {"$exists": True, "$ne": []}}).to_list(1000)
    for u in users_with_favs:
        total_favorites += len(u.get("favorites", []))
    
    # Active providers (is_active=True)
    active_providers = await db.providers.count_documents({"is_active": True})
    inactive_provider_count = await db.providers.count_documents({"is_active": False})
    
    # Average rating across all providers
    pipeline = [
        {"$match": {"average_rating": {"$gt": 0}}},
        {"$group": {"_id": None, "avg": {"$avg": "$average_rating"}}}
    ]
    avg_result = await db.providers.aggregate(pipeline).to_list(1)
    avg_rating = round(avg_result[0]["avg"], 1) if avg_result else 0
    
    # Top 5 providers by rating
    top_providers = await db.providers.find(
        {"total_reviews": {"$gte": 1}},
        {"_id": 0, "name": 1, "average_rating": 1, "total_reviews": 1, "categories": 1}
    ).sort("average_rating", -1).limit(5).to_list(5)
    
    # Recent activity (last 5 providers registered)
    recent_providers = await db.providers.find(
        {},
        {"_id": 0, "name": 1, "created_at": 1, "categories": 1, "cities": 1}
    ).sort("created_at", -1).limit(5).to_list(5)
    
    # Recent reviews (last 5)
    recent_reviews = await db.reviews.find(
        {},
        {"_id": 0, "user_name": 1, "rating": 1, "comment": 1, "created_at": 1, "provider_id": 1}
    ).sort("created_at", -1).limit(5).to_list(5)
    
    # Enrich recent reviews with provider names
    for review in recent_reviews:
        provider = await db.providers.find_one(
            {"provider_id": review.get("provider_id")},
            {"_id": 0, "name": 1}
        )
        review["provider_name"] = provider.get("name", "Desconhecido") if provider else "Desconhecido"
    
    # Category distribution
    cat_pipeline = [
        {"$unwind": "$categories"},
        {"$group": {"_id": "$categories", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 8}
    ]
    category_dist = await db.providers.aggregate(cat_pipeline).to_list(8)
    
    # City distribution
    city_pipeline = [
        {"$unwind": "$cities"},
        {"$group": {"_id": "$cities", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 5}
    ]
    city_dist = await db.providers.aggregate(city_pipeline).to_list(5)
    
    # Users with push tokens (notification reach)
    users_with_tokens = await db.users.count_documents({"push_token": {"$exists": True, "$ne": None}})
    providers_with_tokens = await db.providers.count_documents({"push_token": {"$exists": True, "$ne": None}})
    
    return {
        "total_users": total_users,
        "total_providers": total_providers,
        "active_subscriptions": active_subscriptions,
        "pending_subscriptions": pending_subscriptions,
        "expired_subscriptions": expired_subscriptions,
        "total_reviews": total_reviews,
        "pending_reports": pending_reports,
        "total_reports": total_reports,
        "total_contacts": total_contacts,
        "total_favorites": total_favorites,
        "active_providers": active_providers,
        "inactive_providers": inactive_provider_count,
        "avg_rating": avg_rating,
        "top_providers": top_providers,
        "recent_providers": recent_providers,
        "recent_reviews": recent_reviews,
        "category_distribution": [{"name": c["_id"], "count": c["count"]} for c in category_dist],
        "city_distribution": [{"name": c["_id"], "count": c["count"]} for c in city_dist],
        "notification_reach": {
            "users": users_with_tokens,
            "providers": providers_with_tokens
        }
    }

@api_router.get("/admin/export-excel")
async def export_excel_report(request: Request):
    """Generate and download Excel report with all providers data"""
    await require_admin(request)
    
    # Fetch all data
    providers = await db.providers.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    users = await db.users.find({}, {"_id": 0}).to_list(500)
    subscriptions = await db.subscriptions.find({}, {"_id": 0}).to_list(500)
    
    # Create user email map
    user_email_map = {u.get("user_id"): u.get("email") for u in users}
    
    # Create subscription map
    sub_map = {s.get("provider_id"): s for s in subscriptions}
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    active_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    pending_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    inactive_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== SHEET 1: TODOS OS PRESTADORES ====================
    ws1 = wb.active
    ws1.title = "Todos os Prestadores"
    
    headers1 = ["Nome", "Email", "Telefone", "Categorias", "Cidades", "Bairro", "Status", "Data Validade", "Data Cadastro"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, provider in enumerate(providers, 2):
        email = user_email_map.get(provider.get("user_id"), "N/A")
        categories = ", ".join(provider.get("categories", []))
        cities = ", ".join([c.replace("_", " ").title() for c in provider.get("cities", [])])
        status = provider.get("subscription_status", "inactive")
        
        expires_at = provider.get("subscription_expires_at")
        if expires_at:
            if isinstance(expires_at, str):
                expires_str = expires_at[:10]
            else:
                expires_str = expires_at.strftime("%d/%m/%Y")
        else:
            expires_str = "N/A"
        
        created_at = provider.get("created_at")
        if created_at:
            if isinstance(created_at, str):
                created_str = created_at[:10]
            else:
                created_str = created_at.strftime("%d/%m/%Y")
        else:
            created_str = "N/A"
        
        row_data = [
            provider.get("name", ""),
            email,
            provider.get("phone", ""),
            categories,
            cities,
            provider.get("neighborhood", ""),
            status.upper(),
            expires_str,
            created_str
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            # Color by status
            if status == "active":
                cell.fill = active_fill
            elif status == "pending":
                cell.fill = pending_fill
            elif status in ["inactive", "expired"]:
                cell.fill = inactive_fill
    
    # Adjust column widths
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== SHEET 2: ATIVOS ====================
    ws2 = wb.create_sheet("Ativos")
    active_providers = [p for p in providers if p.get("subscription_status") == "active"]
    
    headers2 = ["Nome", "Email", "Telefone", "WhatsApp", "Categorias", "Cidades", "Data Validade"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, provider in enumerate(active_providers, 2):
        email = user_email_map.get(provider.get("user_id"), "N/A")
        phone = provider.get("phone", "")
        whatsapp = f"https://wa.me/55{phone}" if phone else ""
        categories = ", ".join(provider.get("categories", []))
        cities = ", ".join([c.replace("_", " ").title() for c in provider.get("cities", [])])
        
        expires_at = provider.get("subscription_expires_at")
        if expires_at:
            if isinstance(expires_at, str):
                expires_str = expires_at[:10]
            else:
                expires_str = expires_at.strftime("%d/%m/%Y")
        else:
            expires_str = "N/A"
        
        row_data = [provider.get("name", ""), email, phone, whatsapp, categories, cities, expires_str]
        
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.fill = active_fill
    
    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20
    
    # ==================== SHEET 3: PENDENTES/INATIVOS ====================
    ws3 = wb.create_sheet("Pendentes e Inativos")
    pending_providers = [p for p in providers if p.get("subscription_status") in ["pending", "inactive", None, "expired"]]
    
    headers3 = ["Nome", "Email", "Telefone", "Status", "Categorias", "Cidades", "Data Cadastro"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, provider in enumerate(pending_providers, 2):
        email = user_email_map.get(provider.get("user_id"), "N/A")
        status = provider.get("subscription_status", "inactive")
        categories = ", ".join(provider.get("categories", []))
        cities = ", ".join([c.replace("_", " ").title() for c in provider.get("cities", [])])
        
        created_at = provider.get("created_at")
        if created_at:
            if isinstance(created_at, str):
                created_str = created_at[:10]
            else:
                created_str = created_at.strftime("%d/%m/%Y")
        else:
            created_str = "N/A"
        
        row_data = [provider.get("name", ""), email, provider.get("phone", ""), status.upper(), categories, cities, created_str]
        
        for col, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if status == "pending":
                cell.fill = pending_fill
            else:
                cell.fill = inactive_fill
    
    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    
    # ==================== SHEET 4: RESUMO ====================
    ws4 = wb.create_sheet("Resumo")
    
    total_providers = len(providers)
    total_active = len([p for p in providers if p.get("subscription_status") == "active"])
    total_pending = len([p for p in providers if p.get("subscription_status") == "pending"])
    total_inactive = len([p for p in providers if p.get("subscription_status") in ["inactive", None]])
    total_expired = len([p for p in providers if p.get("subscription_status") == "expired"])
    
    summary_data = [
        ["RESUMO GERAL", ""],
        ["", ""],
        ["Total de Prestadores", total_providers],
        ["Assinaturas Ativas", total_active],
        ["Assinaturas Pendentes", total_pending],
        ["Inativos (sem assinatura)", total_inactive],
        ["Expirados", total_expired],
        ["", ""],
        ["Receita Mensal Potencial", f"R$ {total_active * 15:.2f}"],
        ["", ""],
        ["Data do Relatório", datetime.now().strftime("%d/%m/%Y %H:%M")]
    ]
    
    for row, (label, value) in enumerate(summary_data, 1):
        cell1 = ws4.cell(row=row, column=1, value=label)
        cell2 = ws4.cell(row=row, column=2, value=value)
        
        if row == 1:
            cell1.font = Font(bold=True, size=14)
        elif label and value != "":
            cell1.font = Font(bold=True)
    
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with date
    filename = f"relatorio_achaservico_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/admin/all-providers")
async def get_all_providers(request: Request):
    """Get all providers for admin"""
    await require_admin(request)
    providers = await db.providers.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return providers

@api_router.get("/admin/all-users")
async def get_all_users(request: Request):
    """Get all users for admin"""
    await require_admin(request)
    users = await db.users.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return users

@api_router.get("/admin/all-subscriptions")
async def get_all_subscriptions(request: Request):
    """Get all subscriptions with provider details - automatically filters out orphans"""
    await require_admin(request)
    subscriptions = await db.subscriptions.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # Batch fetch all providers to avoid N+1 query
    provider_ids = [sub.get("provider_id") for sub in subscriptions if sub.get("provider_id")]
    providers = await db.providers.find({"provider_id": {"$in": provider_ids}}, {"_id": 0}).to_list(500)
    provider_map = {p["provider_id"]: p for p in providers}
    
    result = []
    orphan_ids = []
    
    for sub in subscriptions:
        provider = provider_map.get(sub.get("provider_id"))
        if provider:
            # Only include subscriptions with valid providers
            result.append({
                "subscription": sub,
                "provider": provider
            })
        else:
            # Track orphan subscriptions for cleanup
            orphan_ids.append(sub.get("provider_id"))
    
    # Auto-cleanup orphan subscriptions in background
    if orphan_ids:
        for orphan_id in orphan_ids:
            await db.subscriptions.delete_one({"provider_id": orphan_id})
        logger.info(f"Auto-cleaned {len(orphan_ids)} orphan subscriptions")
    
    return result

@api_router.get("/admin/all-reviews")
async def get_all_reviews(request: Request):
    """Get all reviews for admin"""
    await require_admin(request)
    reviews = await db.reviews.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return reviews

@api_router.get("/admin/pending-subscriptions")
async def get_pending_subscriptions(request: Request):
    """Get all pending subscriptions (for admin to activate)"""
    await require_admin(request)
    pending = await db.subscriptions.find({"status": "pending"}, {"_id": 0}).to_list(100)
    
    # Batch fetch all providers to avoid N+1 query
    provider_ids = [sub.get("provider_id") for sub in pending if sub.get("provider_id")]
    providers = await db.providers.find({"provider_id": {"$in": provider_ids}}, {"_id": 0}).to_list(100)
    provider_map = {p["provider_id"]: p for p in providers}
    
    result = []
    for sub in pending:
        provider = provider_map.get(sub.get("provider_id"))
        if provider:
            result.append({
                "subscription": sub,
                "provider": provider
            })
    
    return result

@api_router.post("/admin/activate/{provider_id}")
async def admin_activate_subscription(request: Request, provider_id: str):
    """Admin endpoint to activate a subscription after PIX payment"""
    await require_admin(request)
    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    expires_at = datetime.now(timezone.utc) + timedelta(days=30)
    now = datetime.now(timezone.utc)
    
    # Create or update subscription with full document
    await db.subscriptions.update_one(
        {"provider_id": provider_id},
        {"$set": {
            "provider_id": provider_id,
            "user_id": provider.get("user_id"),
            "status": "active",
            "amount": 9.99,
            "started_at": now,
            "expires_at": expires_at,
            "created_at": now,
            "updated_at": now
        }},
        upsert=True
    )
    
    # Update provider status
    await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": {
            "subscription_status": "active",
            "subscription_expires_at": expires_at,
            "is_active": True
        }}
    )
    
    logger.info(f"Subscription activated for provider: {provider_id} - {provider['name']}")
    
    return {
        "success": True, 
        "message": f"Assinatura de {provider['name']} ativada com sucesso!",
        "expires_at": expires_at.isoformat()
    }

@api_router.post("/admin/cancel-subscription/{provider_id}")
async def admin_cancel_subscription(request: Request, provider_id: str):
    """Cancel a subscription"""
    await require_admin(request)
    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    await db.subscriptions.update_one(
        {"provider_id": provider_id},
        {"$set": {"status": "cancelled"}}
    )
    
    await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": {
            "subscription_status": "inactive",
            "is_active": False
        }}
    )
    
    return {"success": True, "message": "Assinatura cancelada"}

@api_router.post("/admin/check-expired")
async def admin_check_expired_subscriptions(request: Request):
    """Manually check and expire overdue subscriptions"""
    await require_admin(request)
    expired_count = await check_and_expire_subscriptions()
    
    # Get list of expired providers
    expired_providers = await db.providers.find(
        {"subscription_status": "expired"},
        {"_id": 0, "provider_id": 1, "name": 1, "subscription_expires_at": 1}
    ).to_list(100)
    
    return {
        "success": True,
        "expired_count": expired_count,
        "message": f"{expired_count} assinatura(s) expirada(s) foram desativadas",
        "expired_providers": expired_providers
    }

@api_router.get("/admin/expired-subscriptions")
async def get_expired_subscriptions(request: Request):
    """Get all expired subscriptions that can be renewed"""
    await require_admin(request)
    expired = await db.providers.find(
        {"subscription_status": "expired"},
        {"_id": 0}
    ).to_list(100)
    
    # Batch fetch subscriptions to avoid N+1 query
    provider_ids = [p["provider_id"] for p in expired]
    subscriptions = await db.subscriptions.find(
        {"provider_id": {"$in": provider_ids}},
        {"_id": 0}
    ).to_list(100)
    subscription_map = {s["provider_id"]: s for s in subscriptions}
    
    result = []
    for provider in expired:
        subscription = subscription_map.get(provider["provider_id"])
        result.append({
            "provider": provider,
            "subscription": subscription
        })
    
    return result

@api_router.post("/admin/toggle-provider/{provider_id}")
async def admin_toggle_provider(request: Request, provider_id: str):
    """Toggle provider active status"""
    await require_admin(request)
    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    new_status = not provider.get("is_active", True)
    
    await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": {"is_active": new_status}}
    )
    
    return {"success": True, "is_active": new_status}

@api_router.delete("/admin/provider/{provider_id}")
async def admin_delete_provider(request: Request, provider_id: str):
    """Delete a provider"""
    await require_admin(request)
    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    await db.providers.delete_one({"provider_id": provider_id})
    await db.subscriptions.delete_many({"provider_id": provider_id})
    await db.reviews.delete_many({"provider_id": provider_id})
    
    if provider.get("user_id"):
        await db.users.update_one(
            {"user_id": provider["user_id"]},
            {"$set": {"is_provider": False}}
        )
    
    return {"success": True, "message": "Prestador excluído"}

@api_router.delete("/admin/cleanup-orphan-subscriptions")
async def admin_cleanup_orphan_subscriptions(request: Request):
    """Delete subscriptions that don't have a corresponding provider"""
    await require_admin(request)
    # Get all subscriptions
    all_subscriptions = await db.subscriptions.find({}, {"_id": 0, "provider_id": 1}).to_list(1000)
    
    # Get all provider IDs
    all_providers = await db.providers.find({}, {"_id": 0, "provider_id": 1}).to_list(1000)
    provider_ids = {p["provider_id"] for p in all_providers}
    
    # Find orphan subscriptions
    orphan_count = 0
    for sub in all_subscriptions:
        if sub.get("provider_id") not in provider_ids:
            await db.subscriptions.delete_one({"provider_id": sub["provider_id"]})
            orphan_count += 1
    
    return {"success": True, "deleted_count": orphan_count, "message": f"{orphan_count} assinaturas órfãs removidas"}

@api_router.delete("/admin/user/{user_id}")
async def admin_delete_user(request: Request, user_id: str):
    """Delete a user and their provider profile if exists"""
    await require_admin(request)
    user = await db.users.find_one({"user_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    provider = await db.providers.find_one({"user_id": user_id})
    if provider:
        await db.providers.delete_one({"user_id": user_id})
        await db.subscriptions.delete_many({"provider_id": provider["provider_id"]})
        await db.reviews.delete_many({"provider_id": provider["provider_id"]})
    
    await db.reviews.delete_many({"user_id": user_id})
    await db.users.delete_one({"user_id": user_id})
    await db.user_sessions.delete_many({"user_id": user_id})
    
    return {"success": True, "message": "Usuário excluído"}

@api_router.delete("/admin/review/{review_id}")
async def admin_delete_review(request: Request, review_id: str):
    """Delete a review"""
    await require_admin(request)
    review = await db.reviews.find_one({"review_id": review_id})
    if not review:
        raise HTTPException(status_code=404, detail="Avaliação não encontrada")
    
    await db.reviews.delete_one({"review_id": review_id})
    
    # Update provider rating using aggregation (optimized)
    provider_id = review.get("provider_id")
    if provider_id:
        pipeline = [
            {"$match": {"provider_id": provider_id}},
            {"$group": {
                "_id": None,
                "avg_rating": {"$avg": "$rating"},
                "total_reviews": {"$sum": 1}
            }}
        ]
        result = await db.reviews.aggregate(pipeline).to_list(1)
        
        if result:
            avg_rating = result[0]["avg_rating"]
            total_reviews = result[0]["total_reviews"]
            await db.providers.update_one(
                {"provider_id": provider_id},
                {"$set": {
                    "average_rating": round(avg_rating, 1),
                    "total_reviews": total_reviews
                }}
            )
        else:
            await db.providers.update_one(
                {"provider_id": provider_id},
                {"$set": {"average_rating": 0, "total_reviews": 0}}
            )
    
    return {"success": True, "message": "Avaliação excluída"}

# ======================== ADMIN BANNER ========================

@api_router.post("/admin/banner")
async def update_banner(request: Request):
    """Update the app banner (admin only)"""
    await require_admin(request)
    data = await request.json()
    image = data.get("image")
    link = data.get("link", "")
    
    if not image:
        raise HTTPException(status_code=400, detail="Imagem é obrigatória")
    
    # Upload to Cloudinary if base64
    image_url = image
    if image.startswith("data:image"):
        image_url = await upload_to_cloudinary(image, "achaservico/banners")
        if not image_url:
            raise HTTPException(status_code=500, detail="Erro ao fazer upload da imagem")
    
    await db.settings.update_one(
        {"key": "app_banner"},
        {"$set": {
            "key": "app_banner",
            "image": image_url,
            "link": link,
            "active": True,
            "updated_at": datetime.now(timezone.utc)
        }},
        upsert=True
    )
    
    return {"success": True, "message": "Banner atualizado com sucesso"}

@api_router.get("/banner")
async def get_banner():
    """Get current active banner (public)"""
    banner = await db.settings.find_one(
        {"key": "app_banner", "active": True},
        {"_id": 0}
    )
    if not banner:
        return {"active": False}
    return banner

@api_router.delete("/admin/banner")
async def delete_banner(request: Request):
    """Remove the app banner"""
    await require_admin(request)
    await db.settings.update_one(
        {"key": "app_banner"},
        {"$set": {"active": False}}
    )
    return {"success": True, "message": "Banner removido"}

# ======================== REPORTS / DENÚNCIAS ========================

@api_router.post("/reports")
@limiter.limit("5/minute")
async def create_report(report_data: ReportCreate, request: Request):
    """Create a report against a provider"""
    # Get current user if authenticated
    user = await get_current_user(request)
    
    # Check if provider exists
    provider = await db.providers.find_one({"provider_id": report_data.provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    # Check if user already reported this provider (prevent spam)
    if user:
        existing = await db.reports.find_one({
            "provider_id": report_data.provider_id,
            "reporter_user_id": user.user_id,
            "status": "pending"
        })
        if existing:
            raise HTTPException(status_code=400, detail="Você já denunciou este prestador. Aguarde a análise.")
    
    report = Report(
        provider_id=report_data.provider_id,
        provider_name=provider.get("name", ""),
        reporter_user_id=user.user_id if user else None,
        reporter_email=user.email if user else None,
        reason=report_data.reason,
        description=report_data.description,
    )
    
    await db.reports.insert_one(report.dict())
    
    logger.info(f"New report created: {report.report_id} for provider {report_data.provider_id}")
    
    # Send notification to admin
    try:
        # Send push notification to admin
        admin_user = await db.users.find_one({"email": "israel.moraes03@gmail.com"})
        if admin_user and admin_user.get("push_token"):
            reason_labels = {
                "inappropriate_content": "Conteúdo inadequado",
                "false_info": "Informações falsas",
                "bad_behavior": "Comportamento impróprio",
                "spam": "Spam ou propaganda",
                "other": "Outro motivo",
            }
            reason_text = reason_labels.get(report_data.reason, report_data.reason)
            await send_push_notification(
                admin_user["push_token"],
                "🚨 Nova Denúncia Recebida",
                f"Prestador: {provider.get('name', 'Desconhecido')}\nMotivo: {reason_text}",
                {"type": "report", "report_id": report.report_id}
            )
        
        # Send email to admin
        if RESEND_API_KEY:
            reason_labels = {
                "inappropriate_content": "Conteúdo inadequado",
                "false_info": "Informações falsas",
                "bad_behavior": "Comportamento impróprio",
                "spam": "Spam ou propaganda",
                "other": "Outro motivo",
            }
            reason_text = reason_labels.get(report_data.reason, report_data.reason)
            html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="background: linear-gradient(135deg, #EF4444, #DC2626); padding: 20px; border-radius: 10px 10px 0 0;">
                    <h1 style="color: white; margin: 0; font-size: 24px;">🚨 Nova Denúncia Recebida!</h1>
                </div>
                <div style="background: #f9f9f9; padding: 20px; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 10px 10px;">
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; font-weight: bold; color: #666;">Prestador:</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; color: #333;">{provider.get('name', 'Desconhecido')}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; font-weight: bold; color: #666;">Motivo:</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; color: #EF4444; font-weight: bold;">{reason_text}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; font-weight: bold; color: #666;">Descrição:</td>
                            <td style="padding: 10px 0; border-bottom: 1px solid #e0e0e0; color: #333;">{report_data.description or 'Sem descrição'}</td>
                        </tr>
                        <tr>
                            <td style="padding: 10px 0; font-weight: bold; color: #666;">Denunciado por:</td>
                            <td style="padding: 10px 0; color: #333;">{user.email if user else 'Anônimo'}</td>
                        </tr>
                    </table>
                    <div style="margin-top: 20px; padding: 15px; background: #fef2f2; border-radius: 5px; border-left: 4px solid #EF4444;">
                        <p style="margin: 0; color: #991b1b;">
                            <strong>⚠️ Ação necessária:</strong> Acesse o painel admin do AchaServiço para analisar esta denúncia.
                        </p>
                    </div>
                </div>
            </div>
            """
            params = {
                "from": SENDER_EMAIL,
                "to": [ADMIN_NOTIFICATION_EMAIL],
                "subject": f"🚨 Denúncia: {provider.get('name', 'Desconhecido')} - {reason_text}",
                "html": html_content
            }
            await asyncio.to_thread(resend.Emails.send, params)
            logger.info(f"Admin notification email sent for report {report.report_id}")
    except Exception as e:
        logger.error(f"Failed to send admin notification for report: {str(e)}")
    
    return {"success": True, "message": "Denúncia enviada com sucesso. Iremos analisar."}

@api_router.get("/admin/reports")
async def get_admin_reports(request: Request):
    """Get all reports for admin"""
    await require_admin(request)
    reports = await db.reports.find(
        {},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    return reports

@api_router.put("/admin/reports/{report_id}/accept")
async def admin_accept_report(request: Request, report_id: str):
    """Accept a report - marks as accepted and blocks the provider"""
    await require_admin(request)
    report = await db.reports.find_one({"report_id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Denúncia não encontrada")
    
    # Block the provider
    provider = await db.providers.find_one({"provider_id": report.get("provider_id")})
    if provider:
        await db.providers.update_one(
            {"provider_id": report.get("provider_id")},
            {"$set": {
                "blocked": True,
                "is_active": False,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        # Log to block history
        await db.block_history.insert_one({
            "target_type": "provider",
            "target_id": report.get("provider_id"),
            "target_name": provider.get("name", ""),
            "action": "block",
            "reason": report.get("reason", ""),
            "report_id": report_id,
            "admin_action": True,
            "created_at": datetime.now(timezone.utc)
        })
        
        # Send push notification to blocked provider
        push_token = provider.get("push_token")
        if push_token:
            try:
                await send_push_notification(
                    push_token,
                    "⚠️ Perfil Bloqueado",
                    "Seu perfil foi bloqueado devido a uma denúncia. Entre em contato com o suporte para mais informações.",
                    {"type": "blocked"}
                )
            except Exception as e:
                logger.error(f"Failed to send block notification: {str(e)}")
    
    await db.reports.update_one(
        {"report_id": report_id},
        {"$set": {
            "status": "accepted",
            "resolved_at": datetime.now(timezone.utc)
        }}
    )
    
    logger.info(f"Report {report_id} accepted, provider {report.get('provider_id')} blocked")
    return {"success": True, "message": "Denúncia aceita e prestador bloqueado"}

@api_router.put("/admin/reports/{report_id}/discard")
async def admin_discard_report(request: Request, report_id: str):
    """Discard a report"""
    await require_admin(request)
    report = await db.reports.find_one({"report_id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Denúncia não encontrada")
    
    await db.reports.update_one(
        {"report_id": report_id},
        {"$set": {
            "status": "discarded",
            "resolved_at": datetime.now(timezone.utc)
        }}
    )
    
    logger.info(f"Report {report_id} discarded")
    return {"success": True, "message": "Denúncia descartada"}

@api_router.delete("/admin/reports/{report_id}")
async def admin_delete_report(request: Request, report_id: str):
    """Permanently delete a report"""
    await require_admin(request)
    result = await db.reports.delete_one({"report_id": report_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Denúncia não encontrada")
    
    logger.info(f"Report {report_id} permanently deleted")
    return {"success": True, "message": "Denúncia excluída permanentemente"}

@api_router.post("/admin/providers/{provider_id}/unblock")
async def admin_unblock_provider(request: Request, provider_id: str):
    """Unblock a provider"""
    await require_admin(request)
    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": {
            "blocked": False,
            "is_active": True,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Log to block history
    await db.block_history.insert_one({
        "target_type": "provider",
        "target_id": provider_id,
        "target_name": provider.get("name", ""),
        "action": "unblock",
        "admin_action": True,
        "created_at": datetime.now(timezone.utc)
    })
    
    # Notify provider
    push_token = provider.get("push_token")
    if push_token:
        try:
            await send_push_notification(
                push_token,
                "✅ Perfil Desbloqueado",
                "Seu perfil foi desbloqueado e está visível novamente na plataforma.",
                {"type": "unblocked"}
            )
        except Exception as e:
            logger.error(f"Failed to send unblock notification: {str(e)}")
    
    return {"success": True, "message": "Prestador desbloqueado com sucesso"}

@api_router.post("/admin/users/{user_id}/block")
async def admin_block_user(request: Request, user_id: str):
    """Block a user from accessing the app"""
    await require_admin(request)
    user = await db.users.find_one({"user_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"blocked": True, "updated_at": datetime.now(timezone.utc)}}
    )
    
    # Log to block history
    await db.block_history.insert_one({
        "target_type": "user",
        "target_id": user_id,
        "target_name": user.get("name", ""),
        "target_email": user.get("email", ""),
        "action": "block",
        "admin_action": True,
        "created_at": datetime.now(timezone.utc)
    })
    
    return {"success": True, "message": "Usuário bloqueado"}

@api_router.post("/admin/users/{user_id}/unblock")
async def admin_unblock_user(request: Request, user_id: str):
    """Unblock a user"""
    await require_admin(request)
    user = await db.users.find_one({"user_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"blocked": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    await db.block_history.insert_one({
        "target_type": "user",
        "target_id": user_id,
        "target_name": user.get("name", ""),
        "target_email": user.get("email", ""),
        "action": "unblock",
        "admin_action": True,
        "created_at": datetime.now(timezone.utc)
    })
    
    return {"success": True, "message": "Usuário desbloqueado"}

@api_router.get("/admin/block-history")
async def get_block_history(request: Request):
    """Get block/unblock history"""
    await require_admin(request)
    history = await db.block_history.find(
        {}, {"_id": 0}
    ).sort("created_at", -1).limit(50).to_list(50)
    return history

# ======================== ADMIN PREMIUM & NOTIFICATIONS ========================

@api_router.post("/admin/toggle-premium/{provider_id}")
async def admin_toggle_premium(request: Request, provider_id: str):
    """Toggle premium status for a provider (admin only)"""
    await require_admin(request)
    provider = await db.providers.find_one({"provider_id": provider_id})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    current_premium = provider.get("is_premium", False)
    new_premium = not current_premium
    
    update_data = {"is_premium": new_premium}
    
    # If making premium, also activate subscription with no expiration
    if new_premium:
        update_data["subscription_status"] = "active"
        update_data["is_active"] = True
        update_data["subscription_expires_at"] = None  # Premium never expires
    
    await db.providers.update_one(
        {"provider_id": provider_id},
        {"$set": update_data}
    )
    
    logger.info(f"Provider {provider_id} premium status changed to: {new_premium}")
    return {
        "success": True,
        "is_premium": new_premium,
        "message": f"Prestador {'marcado como Premium' if new_premium else 'removido do Premium'}"
    }

@api_router.get("/admin/premium-providers")
async def admin_get_premium_providers(request: Request):
    """Get all premium providers"""
    await require_admin(request)
    providers = await db.providers.find(
        {"is_premium": True},
        {"_id": 0, "profile_image": 0, "service_photos": 0}
    ).to_list(100)
    return providers

class BroadcastNotification(BaseModel):
    title: str
    message: str
    target: str = "providers"  # "providers", "users", "all"

@api_router.post("/admin/broadcast-notification")
async def admin_broadcast_notification(request: Request, notification: BroadcastNotification):
    """Send push notification to providers, users, or all - with target selection"""
    await require_admin(request)
    
    sent_count = 0
    failed_count = 0
    total_recipients = 0
    
    # Get providers if target is "providers" or "all"
    if notification.target in ["providers", "all"]:
        providers = await db.providers.find(
            {},
            {"_id": 0, "push_token": 1, "name": 1, "provider_id": 1, "user_id": 1}
        ).to_list(1000)
        
        for provider in providers:
            total_recipients += 1
            # Save notification to database
            notif = Notification(
                user_id=provider.get("user_id"),
                title=notification.title,
                message=notification.message,
                notification_type="broadcast"
            )
            await db.notifications.insert_one(notif.model_dump())
            
            # Send push notification if token exists
            push_token = provider.get("push_token")
            if push_token:
                try:
                    success = await send_push_notification(
                        push_token=push_token,
                        title=notification.title,
                        body=notification.message,
                        data={"type": "broadcast", "notification_id": notif.notification_id}
                    )
                    if success:
                        sent_count += 1
                    else:
                        failed_count += 1
                except Exception as e:
                    logger.error(f"Failed to send notification to provider {provider.get('name')}: {e}")
                    failed_count += 1
    
    # Get users (non-providers) if target is "users" or "all"
    if notification.target in ["users", "all"]:
        # Get users that are NOT providers
        if notification.target == "users":
            users = await db.users.find(
                {"is_provider": {"$ne": True}},
                {"_id": 0, "push_token": 1, "name": 1, "user_id": 1}
            ).to_list(1000)
        else:
            # For "all", get users that are NOT already in providers list
            provider_user_ids = [p.get("user_id") for p in providers] if notification.target == "all" else []
            users = await db.users.find(
                {"user_id": {"$nin": provider_user_ids}},
                {"_id": 0, "push_token": 1, "name": 1, "user_id": 1}
            ).to_list(1000)
        
        for user in users:
            total_recipients += 1
            # Save notification to database
            notif = Notification(
                user_id=user.get("user_id"),
                title=notification.title,
                message=notification.message,
                notification_type="broadcast"
            )
            await db.notifications.insert_one(notif.model_dump())
            
            # Send push notification if token exists
            push_token = user.get("push_token")
            if push_token:
                try:
                    success = await send_push_notification(
                        push_token=push_token,
                        title=notification.title,
                        body=notification.message,
                        data={"type": "broadcast", "notification_id": notif.notification_id}
                    )
                    if success:
                        sent_count += 1
                    else:
                        failed_count += 1
                except Exception as e:
                    logger.error(f"Failed to send notification to user {user.get('name')}: {e}")
                    failed_count += 1
    
    target_label = {
        "providers": "prestadores",
        "users": "clientes",
        "all": "todos"
    }.get(notification.target, notification.target)
    
    logger.info(f"Broadcast notification ({target_label}) sent: {sent_count} success, {failed_count} failed")
    
    return {
        "success": True,
        "message": f"Notificação enviada para {total_recipients} {target_label}, {sent_count} push enviados",
        "sent": sent_count,
        "failed": failed_count,
        "total_recipients": total_recipients,
        "target": notification.target
    }

# ======================== NOTIFICATIONS ENDPOINTS ========================

@api_router.get("/notifications")
async def get_notifications(request: Request):
    """Get user's notifications"""
    user = await require_auth(request)
    
    notifications = await db.notifications.find(
        {"user_id": user.user_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    return notifications

@api_router.get("/notifications/unread-count")
async def get_unread_count(request: Request):
    """Get count of unread notifications"""
    user = await require_auth(request)
    
    count = await db.notifications.count_documents({
        "user_id": user.user_id,
        "is_read": False
    })
    
    return {"count": count}

@api_router.post("/notifications/mark-read")
async def mark_notifications_read(request: Request):
    """Mark all notifications as read"""
    user = await require_auth(request)
    
    await db.notifications.update_many(
        {"user_id": user.user_id, "is_read": False},
        {"$set": {"is_read": True}}
    )
    
    return {"success": True}

@api_router.post("/admin/send-expiration-notifications")
async def admin_send_expiration_notifications(request: Request):
    """Manually trigger expiration notification check (admin only)"""
    await require_admin(request)
    notifications_sent = await check_expiring_subscriptions()
    return {
        "success": True,
        "notifications_sent": notifications_sent,
        "message": f"{notifications_sent} notificações enviadas"
    }

@api_router.post("/admin/send-review-reminders")
async def admin_send_review_reminders(request: Request):
    """Send review reminder notifications to users who contacted providers 24h ago"""
    await require_admin(request)
    reminders_sent = await send_review_reminders()
    return {
        "success": True,
        "reminders_sent": reminders_sent,
        "message": f"{reminders_sent} lembretes de avaliação enviados"
    }

@api_router.get("/cron/review-reminders")
async def cron_review_reminders():
    """Cron endpoint for review reminders - can be called by external cron service"""
    reminders_sent = await send_review_reminders()
    return {
        "success": True,
        "reminders_sent": reminders_sent
    }

@api_router.post("/providers/register-push-token")
async def register_push_token(request: Request):
    """Register push notification token for a provider"""
    user = await require_auth(request)
    body = await request.json()
    push_token = body.get("push_token")
    
    if not push_token:
        raise HTTPException(status_code=400, detail="Push token é obrigatório")
    
    result = await db.providers.update_one(
        {"user_id": user.user_id},
        {"$set": {"push_token": push_token}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    return {"success": True, "message": "Push token registrado"}

@api_router.post("/users/register-push-token")
async def register_user_push_token(request: Request):
    """Register push notification token for a user (non-provider)"""
    user = await require_auth(request)
    body = await request.json()
    push_token = body.get("push_token")
    
    if not push_token:
        raise HTTPException(status_code=400, detail="Push token é obrigatório")
    
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"push_token": push_token}}
    )
    
    return {"success": True, "message": "Push token registrado"}

# ======================== MERCADO PAGO CHECKOUT PRO ========================

@api_router.post("/mercadopago/create-pix")
async def create_mercadopago_checkout(request: Request):
    """Create a Mercado Pago Checkout Pro preference for PIX/Card payment"""
    user = await require_auth(request)
    
    if not mp_sdk:
        raise HTTPException(status_code=500, detail="Mercado Pago não configurado")
    
    # Get provider
    provider = await db.providers.find_one({"user_id": user.user_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=404, detail="Perfil de prestador não encontrado")
    
    # Check if already has active subscription
    if provider.get("subscription_status") == "active":
        raise HTTPException(status_code=400, detail="Você já possui uma assinatura ativa")
    
    provider_id = provider.get("provider_id")
    
    try:
        # Create Checkout Pro preference (supports PIX, Card, etc.)
        preference_data = {
            "items": [
                {
                    "title": "Assinatura Mensal AchaServiço",
                    "description": f"Assinatura para {provider.get('name')}",
                    "quantity": 1,
                    "unit_price": 9.99,
                    "currency_id": "BRL"
                }
            ],
            "payer": {
                "email": user.email,
                "name": provider.get("name", "Cliente")
            },
            "payment_methods": {
                "excluded_payment_types": [],
                "installments": 1
            },
            "back_urls": {
                "success": f"{APP_DOMAIN}/api/mercadopago/callback?status=success&provider_id={provider_id}",
                "failure": f"{APP_DOMAIN}/api/mercadopago/callback?status=failure&provider_id={provider_id}",
                "pending": f"{APP_DOMAIN}/api/mercadopago/callback?status=pending&provider_id={provider_id}"
            },
            "auto_return": "approved",
            "external_reference": provider_id,
            "notification_url": f"{APP_DOMAIN}/api/mercadopago/webhook",
            "statement_descriptor": "ACHASERVICO"
        }
        
        preference_response = mp_sdk.preference().create(preference_data)
        preference = preference_response.get("response", {})
        
        if preference_response.get("status") not in [200, 201]:
            logger.error(f"Mercado Pago preference error: {preference_response}")
            raise HTTPException(status_code=400, detail="Erro ao criar preferência de pagamento")
        
        preference_id = preference.get("id")
        init_point = preference.get("init_point")  # Production URL
        sandbox_init_point = preference.get("sandbox_init_point")  # Sandbox URL
        
        # Store preference reference
        await db.mp_payments.insert_one({
            "preference_id": preference_id,
            "provider_id": provider_id,
            "user_id": user.user_id,
            "status": "pending",
            "amount": 9.99,
            "created_at": datetime.now(timezone.utc)
        })
        
        logger.info(f"Mercado Pago Checkout created for provider: {provider_id} - Preference ID: {preference_id}")
        
        return {
            "success": True,
            "preference_id": preference_id,
            "checkout_url": init_point,  # Use production URL
            "sandbox_url": sandbox_init_point
        }
        
    except Exception as e:
        logger.error(f"Error creating Mercado Pago Checkout: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao criar pagamento: {str(e)}")

@api_router.get("/mercadopago/callback")
async def mercadopago_callback(
    status: str = "unknown",
    provider_id: str = None,
    collection_id: str = None,
    collection_status: str = None,
    payment_id: str = None,
    external_reference: str = None
):
    """Handle Mercado Pago redirect callback after payment"""
    logger.info(f"Mercado Pago callback: status={status}, provider_id={provider_id}, payment_id={payment_id}, collection_status={collection_status}")
    
    # Use external_reference as provider_id if not provided
    if not provider_id and external_reference:
        provider_id = external_reference
    
    if status == "success" or collection_status == "approved":
        # Payment was successful - activate subscription
        if provider_id:
            expires_at = datetime.now(timezone.utc) + timedelta(days=30)
            now = datetime.now(timezone.utc)
            
            # Update subscription
            await db.subscriptions.update_one(
                {"provider_id": provider_id},
                {"$set": {
                    "provider_id": provider_id,
                    "status": "active",
                    "amount": 9.99,
                    "payment_method": "mercadopago",
                    "mp_payment_id": str(payment_id or collection_id),
                    "started_at": now,
                    "expires_at": expires_at,
                    "updated_at": now
                }},
                upsert=True
            )
            
            # Update provider status
            await db.providers.update_one(
                {"provider_id": provider_id},
                {"$set": {
                    "subscription_status": "active",
                    "subscription_expires_at": expires_at,
                    "is_active": True
                }}
            )
            
            # Update stored payment
            await db.mp_payments.update_one(
                {"provider_id": provider_id, "status": "pending"},
                {"$set": {"status": "approved", "payment_id": str(payment_id or collection_id), "approved_at": now}}
            )
            
            logger.info(f"Mercado Pago subscription activated via callback for provider: {provider_id}")
        
        # Redirect to success page or deep link
        return HTMLResponse(content=f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Pagamento Confirmado</title>
            <style>
                body {{ font-family: Arial, sans-serif; background: #0A0A0A; color: white; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; }}
                .container {{ text-align: center; padding: 40px; }}
                .icon {{ font-size: 80px; margin-bottom: 20px; }}
                h1 {{ color: #10B981; margin-bottom: 10px; }}
                p {{ color: #9CA3AF; margin-bottom: 30px; }}
                .btn {{ background: #10B981; color: white; padding: 15px 30px; border-radius: 10px; text-decoration: none; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="icon">✅</div>
                <h1>Pagamento Confirmado!</h1>
                <p>Sua assinatura foi ativada com sucesso.<br>Você já pode voltar para o aplicativo.</p>
                <a href="achaservico://dashboard" class="btn">Voltar para o App</a>
            </div>
        </body>
        </html>
        """, status_code=200)
    
    elif status == "pending" or collection_status == "pending":
        return HTMLResponse(content=f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Pagamento Pendente</title>
            <style>
                body {{ font-family: Arial, sans-serif; background: #0A0A0A; color: white; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; }}
                .container {{ text-align: center; padding: 40px; }}
                .icon {{ font-size: 80px; margin-bottom: 20px; }}
                h1 {{ color: #F59E0B; margin-bottom: 10px; }}
                p {{ color: #9CA3AF; margin-bottom: 30px; }}
                .btn {{ background: #F59E0B; color: white; padding: 15px 30px; border-radius: 10px; text-decoration: none; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="icon">⏳</div>
                <h1>Pagamento Pendente</h1>
                <p>Seu pagamento está sendo processado.<br>Você receberá uma notificação quando for confirmado.</p>
                <a href="achaservico://dashboard" class="btn">Voltar para o App</a>
            </div>
        </body>
        </html>
        """, status_code=200)
    
    else:
        return HTMLResponse(content=f"""
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Pagamento Não Concluído</title>
            <style>
                body {{ font-family: Arial, sans-serif; background: #0A0A0A; color: white; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; }}
                .container {{ text-align: center; padding: 40px; }}
                .icon {{ font-size: 80px; margin-bottom: 20px; }}
                h1 {{ color: #EF4444; margin-bottom: 10px; }}
                p {{ color: #9CA3AF; margin-bottom: 30px; }}
                .btn {{ background: #374151; color: white; padding: 15px 30px; border-radius: 10px; text-decoration: none; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="icon">❌</div>
                <h1>Pagamento Não Concluído</h1>
                <p>O pagamento não foi finalizado.<br>Tente novamente pelo aplicativo.</p>
                <a href="achaservico://dashboard" class="btn">Voltar para o App</a>
            </div>
        </body>
        </html>
        """, status_code=200)

@api_router.get("/mercadopago/payment-status/{payment_id}")
async def get_mercadopago_payment_status(payment_id: str):
    """Check the status of a Mercado Pago payment"""
    if not mp_sdk:
        raise HTTPException(status_code=500, detail="Mercado Pago não configurado")
    
    try:
        payment_response = mp_sdk.payment().get(payment_id)
        payment = payment_response.get("response", {})
        
        return {
            "payment_id": payment.get("id"),
            "status": payment.get("status"),
            "status_detail": payment.get("status_detail"),
            "approved": payment.get("status") == "approved"
        }
    except Exception as e:
        logger.error(f"Error checking Mercado Pago payment: {str(e)}")
        raise HTTPException(status_code=404, detail="Pagamento não encontrado")

@api_router.post("/mercadopago/webhook")
async def mercadopago_webhook(request: Request):
    """Handle Mercado Pago webhooks for payment notifications"""
    try:
        body = await request.json()
        logger.info(f"Mercado Pago webhook received: {body}")
        
        action = body.get("action")
        data = body.get("data", {})
        payment_id = data.get("id")
        
        if action == "payment.updated" or action == "payment.created":
            if payment_id and mp_sdk:
                # Get payment details
                payment_response = mp_sdk.payment().get(payment_id)
                payment = payment_response.get("response", {})
                
                if payment.get("status") == "approved":
                    # Get provider_id from external_reference (Checkout Pro) or metadata
                    provider_id = payment.get("external_reference")
                    
                    if not provider_id:
                        metadata = payment.get("metadata", {})
                        provider_id = metadata.get("provider_id")
                    
                    if not provider_id:
                        # Try to find from stored payment
                        stored_payment = await db.mp_payments.find_one({"payment_id": str(payment_id)})
                        if stored_payment:
                            provider_id = stored_payment.get("provider_id")
                    
                    if provider_id:
                        # Activate subscription
                        expires_at = datetime.now(timezone.utc) + timedelta(days=30)
                        now = datetime.now(timezone.utc)
                        
                        # Create/update subscription
                        await db.subscriptions.update_one(
                            {"provider_id": provider_id},
                            {"$set": {
                                "provider_id": provider_id,
                                "status": "active",
                                "amount": 9.99,
                                "payment_method": "mercadopago_pix",
                                "mp_payment_id": str(payment_id),
                                "started_at": now,
                                "expires_at": expires_at,
                                "created_at": now,
                                "updated_at": now
                            }},
                            upsert=True
                        )
                        
                        # Update provider status
                        await db.providers.update_one(
                            {"provider_id": provider_id},
                            {"$set": {
                                "subscription_status": "active",
                                "subscription_expires_at": expires_at,
                                "is_active": True
                            }}
                        )
                        
                        # Update stored payment
                        await db.mp_payments.update_one(
                            {"payment_id": str(payment_id)},
                            {"$set": {"status": "approved", "approved_at": now}}
                        )
                        
                        logger.info(f"Mercado Pago subscription activated for provider: {provider_id}")
        
        return {"status": "ok"}
        
    except Exception as e:
        logger.error(f"Error processing Mercado Pago webhook: {str(e)}")
        return {"status": "error", "message": str(e)}

@api_router.post("/mercadopago/activate-from-payment")
async def activate_from_mercadopago_payment(request: Request):
    """Manually activate subscription from Mercado Pago payment (fallback)"""
    user = await require_auth(request)
    
    body = await request.json()
    payment_id = body.get("payment_id")
    
    if not payment_id or not mp_sdk:
        raise HTTPException(status_code=400, detail="Payment ID é obrigatório")
    
    try:
        # Get payment from Mercado Pago
        payment_response = mp_sdk.payment().get(payment_id)
        payment = payment_response.get("response", {})
        
        if payment.get("status") != "approved":
            raise HTTPException(status_code=400, detail="Pagamento não foi aprovado")
        
        # Get provider
        provider = await db.providers.find_one({"user_id": user.user_id}, {"_id": 0})
        if not provider:
            raise HTTPException(status_code=404, detail="Prestador não encontrado")
        
        provider_id = provider.get("provider_id")
        
        # Check if already active
        if provider.get("subscription_status") == "active":
            return {"success": True, "message": "Assinatura já está ativa", "already_active": True}
        
        # Activate subscription
        expires_at = datetime.now(timezone.utc) + timedelta(days=30)
        now = datetime.now(timezone.utc)
        
        await db.subscriptions.update_one(
            {"provider_id": provider_id},
            {"$set": {
                "provider_id": provider_id,
                "status": "active",
                "amount": 9.99,
                "payment_method": "mercadopago_pix",
                "mp_payment_id": str(payment_id),
                "started_at": now,
                "expires_at": expires_at,
                "updated_at": now
            }},
            upsert=True
        )
        
        await db.providers.update_one(
            {"provider_id": provider_id},
            {"$set": {
                "subscription_status": "active",
                "subscription_expires_at": expires_at,
                "is_active": True
            }}
        )
        
        logger.info(f"Mercado Pago subscription manually activated for provider: {provider_id}")
        
        return {"success": True, "message": "Assinatura ativada com sucesso!", "expires_at": expires_at.isoformat()}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error activating from Mercado Pago: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro ao ativar assinatura: {str(e)}")


@api_router.post("/mercadopago/check-and-activate")
async def check_and_activate_mercadopago(request: Request):
    """Check if there's an approved payment for the user and activate subscription"""
    user = await require_auth(request)
    
    if not mp_sdk:
        raise HTTPException(status_code=500, detail="Mercado Pago não configurado")
    
    body = await request.json()
    preference_id = body.get("preference_id")
    
    # Get provider
    provider = await db.providers.find_one({"user_id": user.user_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=404, detail="Prestador não encontrado")
    
    provider_id = provider.get("provider_id")
    
    # Check if already active
    if provider.get("subscription_status") == "active":
        return {"activated": True, "message": "Assinatura já está ativa", "already_active": True}
    
    try:
        # Search for approved payments with this external_reference (provider_id)
        search_response = mp_sdk.payment().search({
            "external_reference": provider_id,
            "status": "approved"
        })
        
        payments = search_response.get("response", {}).get("results", [])
        
        if payments:
            # Found approved payment - activate subscription
            payment = payments[0]
            payment_id = payment.get("id")
            
            expires_at = datetime.now(timezone.utc) + timedelta(days=30)
            now = datetime.now(timezone.utc)
            
            await db.subscriptions.update_one(
                {"provider_id": provider_id},
                {"$set": {
                    "provider_id": provider_id,
                    "status": "active",
                    "amount": 9.99,
                    "payment_method": "mercadopago",
                    "mp_payment_id": str(payment_id),
                    "started_at": now,
                    "expires_at": expires_at,
                    "created_at": now,
                    "updated_at": now
                }},
                upsert=True
            )
            
            await db.providers.update_one(
                {"provider_id": provider_id},
                {"$set": {
                    "subscription_status": "active",
                    "subscription_expires_at": expires_at,
                    "is_active": True
                }}
            )
            
            logger.info(f"Mercado Pago subscription activated via check for provider: {provider_id}")
            return {"activated": True, "message": "Assinatura ativada com sucesso!", "expires_at": expires_at.isoformat()}
        
        return {"activated": False, "message": "Nenhum pagamento aprovado encontrado"}
        
    except Exception as e:
        logger.error(f"Error checking Mercado Pago payment: {str(e)}")
        return {"activated": False, "message": str(e)}


# ======================== STRIPE PAYMENTS ========================

@api_router.post("/stripe/create-checkout-session")
async def create_stripe_checkout_session(request: Request):
    """Create a Stripe Checkout session for subscription payment"""
    user = await require_auth(request)
    
    provider = await db.providers.find_one({"user_id": user.user_id}, {"_id": 0})
    if not provider:
        raise HTTPException(status_code=400, detail="Você precisa criar um perfil de prestador primeiro")
    
    # Check if already has active subscription
    if provider.get("subscription_status") == "active":
        raise HTTPException(status_code=400, detail="Você já possui uma assinatura ativa")
    
    try:
        # Create Stripe Checkout Session
        checkout_session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': 'brl',
                    'product_data': {
                        'name': 'Assinatura Mensal AchaServiço',
                        'description': 'Acesso completo para prestadores de serviços em Três Lagoas',
                    },
                    'unit_amount': 9.99,  # R$ 9.99 in centavos
                },
                'quantity': 1,
            }],
            mode='payment',
            # Use backend URL for return, which will redirect to app via JavaScript
            success_url=f"{APP_DOMAIN}/api/stripe/payment-complete?session_id={{CHECKOUT_SESSION_ID}}&status=success",
            cancel_url=f"{APP_DOMAIN}/api/stripe/payment-complete?status=cancelled",
            metadata={
                'provider_id': provider['provider_id'],
                'user_id': user.user_id,
            },
            customer_email=user.email,
        )
        
        logger.info(f"Stripe checkout session created for provider: {provider['provider_id']}")
        
        return {
            "checkout_url": checkout_session.url,
            "session_id": checkout_session.id
        }
        
    except stripe.error.StripeError as e:
        logger.error(f"Stripe error: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao processar pagamento: {str(e)}")

@api_router.post("/stripe/webhook")
async def stripe_webhook(request: Request):
    """Handle Stripe webhook events"""
    payload = await request.body()
    sig_header = request.headers.get('stripe-signature')
    
    # If webhook secret is set, verify signature
    if STRIPE_WEBHOOK_SECRET:
        try:
            event = stripe.Webhook.construct_event(
                payload, sig_header, STRIPE_WEBHOOK_SECRET
            )
        except ValueError as e:
            logger.error(f"Invalid payload: {e}")
            raise HTTPException(status_code=400, detail="Invalid payload")
        except stripe.error.SignatureVerificationError as e:
            logger.error(f"Invalid signature: {e}")
            raise HTTPException(status_code=400, detail="Invalid signature")
    else:
        # No webhook secret - parse JSON directly (for testing)
        try:
            event = stripe.Event.construct_from(
                await request.json(), stripe.api_key
            )
        except Exception as e:
            logger.error(f"Error parsing event: {e}")
            raise HTTPException(status_code=400, detail="Invalid event")
    
    # Store webhook for audit
    await db.webhooks.insert_one({
        "type": "stripe",
        "event_type": event.type,
        "event_id": event.id,
        "data": event.data.object if hasattr(event.data, 'object') else str(event.data),
        "received_at": datetime.now(timezone.utc),
        "processed": False
    })
    
    # Handle checkout.session.completed
    if event.type == 'checkout.session.completed':
        session = event.data.object
        provider_id = session.metadata.get('provider_id')
        user_id = session.metadata.get('user_id')
        
        if provider_id:
            expires_at = datetime.now(timezone.utc) + timedelta(days=30)
            now = datetime.now(timezone.utc)
            
            # Create or update subscription
            await db.subscriptions.update_one(
                {"provider_id": provider_id},
                {"$set": {
                    "provider_id": provider_id,
                    "user_id": user_id,
                    "status": "active",
                    "amount": 9.99,
                    "payment_method": "stripe",
                    "stripe_session_id": session.id,
                    "stripe_payment_intent": session.payment_intent,
                    "started_at": now,
                    "expires_at": expires_at,
                    "created_at": now,
                    "updated_at": now
                }},
                upsert=True
            )
            
            # Activate provider
            await db.providers.update_one(
                {"provider_id": provider_id},
                {"$set": {
                    "subscription_status": "active",
                    "subscription_expires_at": expires_at,
                    "is_active": True
                }}
            )
            
            logger.info(f"Stripe payment completed - Provider {provider_id} activated until {expires_at}")
            
            # Mark webhook as processed
            await db.webhooks.update_one(
                {"event_id": event.id},
                {"$set": {"processed": True}}
            )
    
    elif event.type == 'payment_intent.payment_failed':
        logger.warning(f"Payment failed: {event.data.object}")
    
    return {"status": "ok"}

@api_router.get("/stripe/payment-complete", response_class=HTMLResponse)
async def stripe_payment_complete(session_id: str = None, status: str = "success"):
    """HTML page shown after Stripe payment - instructs user to close browser and return to app"""
    
    if status == "success" and session_id:
        # Try to activate subscription automatically
        try:
            session = stripe.checkout.Session.retrieve(session_id)
            if session.payment_status == 'paid':
                provider_id = session.metadata.get('provider_id')
                if provider_id:
                    provider = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0})
                    if provider and provider.get("subscription_status") != "active":
                        expires_at = datetime.now(timezone.utc) + timedelta(days=30)
                        now = datetime.now(timezone.utc)
                        
                        await db.subscriptions.update_one(
                            {"provider_id": provider_id},
                            {"$set": {
                                "provider_id": provider_id,
                                "user_id": provider.get("user_id"),
                                "status": "active",
                                "amount": 9.99,
                                "payment_method": "stripe",
                                "stripe_session_id": session.id,
                                "started_at": now,
                                "expires_at": expires_at,
                            }},
                            upsert=True
                        )
                        
                        await db.providers.update_one(
                            {"provider_id": provider_id},
                            {"$set": {
                                "subscription_status": "active",
                                "subscription_expires_at": expires_at,
                                "is_active": True
                            }}
                        )
                        logger.info(f"Subscription auto-activated for provider: {provider_id}")
        except Exception as e:
            logger.error(f"Error auto-activating subscription: {e}")
    
    # Return HTML page
    if status == "success":
        html_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Pagamento Confirmado - AchaServiço</title>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background: linear-gradient(135deg, #0A0A0A 0%, #1a1a2e 100%);
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    color: white;
                    padding: 20px;
                }
                .container {
                    text-align: center;
                    max-width: 400px;
                }
                .icon {
                    width: 80px;
                    height: 80px;
                    background: #10B981;
                    border-radius: 50%;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    margin: 0 auto 24px;
                    font-size: 40px;
                }
                h1 {
                    font-size: 28px;
                    margin-bottom: 16px;
                    color: #10B981;
                }
                p {
                    font-size: 16px;
                    color: #9CA3AF;
                    margin-bottom: 32px;
                    line-height: 1.5;
                }
                .btn {
                    background: #10B981;
                    color: white;
                    border: none;
                    padding: 16px 32px;
                    font-size: 18px;
                    border-radius: 12px;
                    cursor: pointer;
                    font-weight: 600;
                    width: 100%;
                    max-width: 280px;
                }
                .btn:active { background: #059669; }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="icon">✓</div>
                <h1>Pagamento Confirmado!</h1>
                <p>Sua assinatura foi ativada com sucesso. Agora você pode receber contatos de clientes.</p>
                <button class="btn" onclick="window.close()">Fechar e Voltar ao App</button>
                <p style="margin-top: 20px; font-size: 14px;">Se o botão não funcionar, feche esta aba manualmente.</p>
            </div>
        </body>
        </html>
        """
    else:
        html_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Pagamento Cancelado - AchaServiço</title>
            <style>
                * { margin: 0; padding: 0; box-sizing: border-box; }
                body {
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background: linear-gradient(135deg, #0A0A0A 0%, #1a1a2e 100%);
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                    color: white;
                    padding: 20px;
                }
                .container { text-align: center; max-width: 400px; }
                .icon {
                    width: 80px; height: 80px;
                    background: #EF4444;
                    border-radius: 50%;
                    display: flex; align-items: center; justify-content: center;
                    margin: 0 auto 24px;
                    font-size: 40px;
                }
                h1 { font-size: 28px; margin-bottom: 16px; color: #EF4444; }
                p { font-size: 16px; color: #9CA3AF; margin-bottom: 32px; }
                .btn {
                    background: #374151; color: white; border: none;
                    padding: 16px 32px; font-size: 18px; border-radius: 12px;
                    cursor: pointer; font-weight: 600;
                }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="icon">✕</div>
                <h1>Pagamento Cancelado</h1>
                <p>O pagamento foi cancelado. Você pode tentar novamente quando quiser.</p>
                <button class="btn" onclick="window.close()">Fechar e Voltar ao App</button>
            </div>
        </body>
        </html>
        """
    
    return HTMLResponse(content=html_content)

@api_router.get("/stripe/payment-status/{session_id}")
async def get_payment_status(session_id: str, request: Request):
    """Check the status of a Stripe payment session"""
    try:
        session = stripe.checkout.Session.retrieve(session_id)
        
        logger.info(f"Payment status check for session {session_id}: status={session.payment_status}")
        
        return {
            "status": session.payment_status,
            "provider_id": session.metadata.get('provider_id'),
            "completed": session.payment_status == 'paid'
        }
    except stripe.error.StripeError as e:
        logger.error(f"Error retrieving session: {e}")
        raise HTTPException(status_code=404, detail="Sessão não encontrada")

class ActivateFromSessionRequest(BaseModel):
    session_id: str

@api_router.post("/stripe/activate-from-session")
async def activate_from_stripe_session(data: ActivateFromSessionRequest, request: Request):
    """Manually activate subscription from Stripe session (fallback when webhook doesn't work)"""
    # Note: Authentication is done via Stripe session verification, not user token
    # This allows activation even if session was lost during Stripe redirect
    
    try:
        # Retrieve the session from Stripe
        session = stripe.checkout.Session.retrieve(data.session_id)
        
        # Verify payment was successful
        if session.payment_status != 'paid':
            raise HTTPException(status_code=400, detail="Pagamento não foi concluído")
        
        # Get provider info from session metadata
        provider_id = session.metadata.get('provider_id')
        session_user_id = session.metadata.get('user_id')
        
        if not provider_id:
            raise HTTPException(status_code=400, detail="Sessão inválida - provider_id não encontrado")
        
        # Get the provider (using provider_id from session metadata for security)
        provider = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0})
        if not provider:
            raise HTTPException(status_code=404, detail="Prestador não encontrado")
        
        # Check if already activated
        if provider.get("subscription_status") == "active":
            return {"success": True, "message": "Assinatura já está ativa", "already_active": True}
        
        # Activate the subscription
        expires_at = datetime.now(timezone.utc) + timedelta(days=30)
        now = datetime.now(timezone.utc)
        
        # Create or update subscription record
        await db.subscriptions.update_one(
            {"provider_id": provider_id},
            {"$set": {
                "provider_id": provider_id,
                "user_id": session_user_id or provider.get("user_id"),
                "status": "active",
                "amount": 9.99,
                "payment_method": "stripe",
                "stripe_session_id": session.id,
                "stripe_payment_intent": session.payment_intent,
                "started_at": now,
                "expires_at": expires_at,
                "created_at": now,
                "updated_at": now
            }},
            upsert=True
        )
        
        # Activate provider
        await db.providers.update_one(
            {"provider_id": provider_id},
            {"$set": {
                "subscription_status": "active",
                "subscription_expires_at": expires_at,
                "is_active": True
            }}
        )
        
        logger.info(f"Stripe subscription manually activated for provider: {provider_id} - {provider.get('name')}")
        
        return {"success": True, "message": "Assinatura ativada com sucesso!", "expires_at": expires_at.isoformat()}
        
    except stripe.error.StripeError as e:
        logger.error(f"Stripe error checking session: {e}")
        raise HTTPException(status_code=400, detail=f"Erro ao verificar pagamento: {str(e)}")

# ======================== WEBHOOKS ========================

@api_router.post("/webhooks/pix")
async def pix_webhook(request: Request):
    """Handle PIX webhook notifications (for future use)"""
    try:
        body = await request.json()
        logger.info(f"PIX Webhook received: {body}")
        
        # Store webhook for audit
        await db.webhooks.insert_one({
            "type": "pix",
            "data": body,
            "received_at": datetime.now(timezone.utc),
            "processed": False
        })
        
        return {"status": "ok"}
        
    except Exception as e:
        logger.error(f"Webhook error: {e}")
        return {"status": "error", "message": str(e)}

# ======================== LEGAL PAGES ========================

PRIVACY_POLICY_HTML = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Política de Privacidade - AchaServiço</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #0A0A0A; color: #E5E5E5; line-height: 1.7; padding: 20px; }
        .container { max-width: 800px; margin: 0 auto; }
        h1 { color: #10B981; font-size: 28px; margin-bottom: 10px; }
        h2 { color: #10B981; font-size: 20px; margin: 30px 0 15px; }
        p { margin-bottom: 15px; }
        ul { margin: 15px 0 15px 25px; }
        li { margin-bottom: 8px; }
        .updated { color: #6B7280; font-size: 14px; margin-bottom: 30px; }
        .highlight { background: #10B98120; padding: 15px; border-radius: 8px; border-left: 4px solid #10B981; margin: 20px 0; }
        a { color: #10B981; }
        .footer { margin-top: 40px; padding-top: 20px; border-top: 1px solid #333; color: #6B7280; font-size: 14px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Política de Privacidade</h1>
        <p class="updated">Última atualização: Março de 2026</p>
        
        <div class="highlight">
            <strong>Resumo:</strong> O AchaServiço coleta apenas as informações necessárias para conectar você a prestadores de serviços. Não vendemos seus dados.
        </div>
        
        <h2>1. Informações que Coletamos</h2>
        <p>Ao utilizar o AchaServiço, podemos coletar:</p>
        <ul>
            <li><strong>Dados de cadastro:</strong> Nome, e-mail e foto de perfil (via login Google)</li>
            <li><strong>Dados de prestadores:</strong> Nome, telefone/WhatsApp, descrição dos serviços, fotos dos trabalhos, cidade e bairros de atuação</li>
            <li><strong>Dados de uso:</strong> Avaliações, favoritos e histórico de contatos</li>
            <li><strong>Imagens e fotos:</strong> Fotos tiradas pela câmera do dispositivo ou selecionadas da galeria para o perfil e portfólio de serviços</li>
        </ul>
        
        <h2>2. Uso da Câmera e Galeria</h2>
        <p>O AchaServiço solicita permissão para acessar:</p>
        <ul>
            <li><strong>Câmera:</strong> Para permitir que prestadores tirem fotos de seus trabalhos e serviços diretamente pelo aplicativo</li>
            <li><strong>Galeria de fotos:</strong> Para permitir a seleção de imagens existentes para o perfil e portfólio</li>
        </ul>
        <p>Essas permissões são utilizadas <strong>exclusivamente</strong> para as funcionalidades do aplicativo. As fotos são armazenadas de forma segura e você pode removê-las a qualquer momento.</p>
        
        <h2>3. Como Usamos suas Informações</h2>
        <p>Utilizamos seus dados para:</p>
        <ul>
            <li>Permitir o cadastro e login no aplicativo</li>
            <li>Exibir perfis de prestadores para usuários que buscam serviços</li>
            <li>Facilitar o contato entre usuários e prestadores via WhatsApp</li>
            <li>Enviar notificações sobre favoritos e comunicados importantes</li>
            <li>Melhorar nossos serviços e experiência do usuário</li>
        </ul>
        
        <h2>4. Compartilhamento de Dados</h2>
        <p>Seus dados podem ser compartilhados:</p>
        <ul>
            <li><strong>Com outros usuários:</strong> Informações públicas do perfil de prestadores são visíveis para quem busca serviços</li>
            <li><strong>Com serviços de terceiros:</strong> Google (autenticação), Cloudinary (armazenamento de imagens), Expo (notificações push)</li>
        </ul>
        <p><strong>Não vendemos, alugamos ou comercializamos seus dados pessoais.</strong></p>
        
        <h2>5. Armazenamento e Segurança</h2>
        <p>Seus dados são armazenados em servidores seguros com criptografia. Utilizamos práticas de segurança padrão da indústria para proteger suas informações.</p>
        
        <h2>6. Seus Direitos</h2>
        <p>Você tem direito a:</p>
        <ul>
            <li>Acessar seus dados pessoais</li>
            <li>Corrigir informações incorretas</li>
            <li>Solicitar a exclusão da sua conta e dados</li>
            <li>Revogar consentimento para uso dos dados</li>
        </ul>
        <p>Para exercer esses direitos, entre em contato conosco.</p>
        
        <h2>7. Cookies e Tecnologias</h2>
        <p>O aplicativo pode utilizar tecnologias de armazenamento local para manter sua sessão ativa e preferências salvas.</p>
        
        <h2>8. Menores de Idade</h2>
        <p>O AchaServiço não é destinado a menores de 18 anos. Não coletamos intencionalmente dados de menores.</p>
        
        <h2>9. Alterações nesta Política</h2>
        <p>Podemos atualizar esta política periodicamente. Notificaremos sobre mudanças significativas através do aplicativo.</p>
        
        <h2>10. Contato</h2>
        <p>Para dúvidas sobre privacidade:</p>
        <ul>
            <li><strong>E-mail:</strong> contato.achaservico@gmail.com</li>
            <li><strong>Desenvolvedor:</strong> Sara Gomes da Silva</li>
            <li><strong>Localização:</strong> Três Lagoas - MS, Brasil</li>
        </ul>
        
        <div class="footer">
            <p>© 2026 AchaServiço. Todos os direitos reservados.</p>
        </div>
    </div>
</body>
</html>
"""

TERMS_OF_USE_HTML = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Termos de Uso - AchaServiço</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #0A0A0A; color: #E5E5E5; line-height: 1.7; padding: 20px; }
        .container { max-width: 800px; margin: 0 auto; }
        h1 { color: #10B981; font-size: 28px; margin-bottom: 10px; }
        h2 { color: #10B981; font-size: 20px; margin: 30px 0 15px; }
        p { margin-bottom: 15px; }
        ul { margin: 15px 0 15px 25px; }
        li { margin-bottom: 8px; }
        .updated { color: #6B7280; font-size: 14px; margin-bottom: 30px; }
        .highlight { background: #10B98120; padding: 15px; border-radius: 8px; border-left: 4px solid #10B981; margin: 20px 0; }
        .warning { background: #EF444420; padding: 15px; border-radius: 8px; border-left: 4px solid #EF4444; margin: 20px 0; }
        a { color: #10B981; }
        .footer { margin-top: 40px; padding-top: 20px; border-top: 1px solid #333; color: #6B7280; font-size: 14px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Termos de Uso</h1>
        <p class="updated">Última atualização: Março de 2026</p>
        
        <div class="highlight">
            <strong>Resumo:</strong> O AchaServiço é uma plataforma que conecta usuários a prestadores de serviços. Não somos responsáveis pelos serviços prestados.
        </div>
        
        <h2>1. Aceitação dos Termos</h2>
        <p>Ao acessar ou usar o AchaServiço, você concorda com estes Termos de Uso. Se não concordar, não utilize o aplicativo.</p>
        
        <h2>2. Descrição do Serviço</h2>
        <p>O AchaServiço é uma plataforma gratuita que:</p>
        <ul>
            <li>Conecta usuários que precisam de serviços a prestadores locais</li>
            <li>Permite que prestadores divulguem seus serviços</li>
            <li>Facilita o contato direto via WhatsApp</li>
            <li>Possibilita avaliações de prestadores</li>
        </ul>
        
        <h2>3. Cadastro e Conta</h2>
        <p>Para utilizar o AchaServiço:</p>
        <ul>
            <li>Você deve ter pelo menos 18 anos</li>
            <li>O cadastro é feito via conta Google</li>
            <li>Você é responsável por manter suas credenciais seguras</li>
            <li>As informações fornecidas devem ser verdadeiras e atualizadas</li>
        </ul>
        
        <h2>4. Regras para Prestadores</h2>
        <p>Prestadores cadastrados devem:</p>
        <ul>
            <li>Fornecer informações verdadeiras sobre seus serviços</li>
            <li>Manter telefone de contato atualizado</li>
            <li>Não publicar conteúdo ofensivo, ilegal ou enganoso</li>
            <li>Respeitar os usuários e responder de forma profissional</li>
            <li>Usar apenas fotos próprias ou com direitos de uso</li>
        </ul>
        
        <h2>5. Regras para Usuários</h2>
        <p>Usuários do AchaServiço devem:</p>
        <ul>
            <li>Usar o aplicativo de forma respeitosa</li>
            <li>Fazer avaliações honestas e construtivas</li>
            <li>Não usar linguagem ofensiva ou discriminatória</li>
            <li>Não tentar fraudar ou manipular o sistema</li>
        </ul>
        
        <div class="warning">
            <h2>6. Isenção de Responsabilidade</h2>
            <p><strong>IMPORTANTE:</strong> O AchaServiço é apenas uma plataforma de conexão.</p>
            <ul>
                <li>Não somos parte dos contratos entre usuários e prestadores</li>
                <li>Não garantimos a qualidade dos serviços prestados</li>
                <li>Não nos responsabilizamos por danos decorrentes dos serviços contratados</li>
                <li>A negociação de preços e condições é feita diretamente entre as partes</li>
            </ul>
        </div>
        
        <h2>7. Avaliações</h2>
        <p>O sistema de avaliações:</p>
        <ul>
            <li>É anônimo para proteger os usuários</li>
            <li>Requer contato prévio com o prestador via WhatsApp</li>
            <li>Deve refletir experiências reais</li>
            <li>Pode ser removido se violar nossas diretrizes</li>
        </ul>
        
        <h2>8. Propriedade Intelectual</h2>
        <p>Todo o conteúdo do AchaServiço (logotipos, design, código) é de propriedade da desenvolvedora. É proibida a reprodução sem autorização.</p>
        
        <h2>9. Suspensão e Encerramento</h2>
        <p>Reservamo-nos o direito de suspender ou encerrar contas que:</p>
        <ul>
            <li>Violem estes termos</li>
            <li>Publiquem conteúdo inadequado</li>
            <li>Prejudiquem outros usuários ou a plataforma</li>
        </ul>
        
        <h2>10. Cancelamento de Conta</h2>
        <p>Você pode solicitar o cancelamento da sua conta a qualquer momento. Para isso:</p>
        <ul>
            <li>Acesse o aplicativo AchaServiço</li>
            <li>Vá até <strong>Perfil → Ajuda / Suporte → Cancelar Conta</strong></li>
            <li>Ao clicar, será aberto seu aplicativo de e-mail com uma solicitação pré-preenchida</li>
            <li>Envie o e-mail para <strong>contato.achaservico@gmail.com</strong></li>
            <li>Nossa equipe processará sua solicitação em até <strong>5 dias úteis</strong></li>
        </ul>
        <p>Ao cancelar sua conta:</p>
        <ul>
            <li>Seus dados pessoais serão removidos permanentemente</li>
            <li>Caso seja prestador, seu perfil será desativado e removido das buscas</li>
            <li>Suas avaliações feitas em outros prestadores serão anonimizadas</li>
            <li>Esta ação é <strong>irreversível</strong></li>
        </ul>
        
        <h2>11. Planos e Preços</h2>
        <p>O AchaServiço é atualmente gratuito para todos os usuários. Reservamo-nos o direito de, no futuro, introduzir planos pagos com recursos adicionais para prestadores de serviços.</p>
        <p>Qualquer alteração nos preços ou introdução de cobranças será comunicada com antecedência mínima de 30 dias através do aplicativo e/ou e-mail cadastrado. O uso continuado após as alterações implica aceitação dos novos termos.</p>
        
        <h2>12. Alterações nos Termos</h2>
        <p>Podemos modificar estes termos a qualquer momento. Mudanças significativas serão comunicadas pelo aplicativo.</p>
        
        <h2>13. Lei Aplicável</h2>
        <p>Estes termos são regidos pelas leis da República Federativa do Brasil.</p>
        
        <h2>14. Contato</h2>
        <p>Para dúvidas sobre estes termos:</p>
        <ul>
            <li><strong>E-mail:</strong> contato.achaservico@gmail.com</li>
            <li><strong>Desenvolvedor:</strong> Sara Gomes da Silva</li>
            <li><strong>Localização:</strong> Três Lagoas - MS, Brasil</li>
        </ul>
        
        <div class="footer">
            <p>© 2026 AchaServiço. Todos os direitos reservados.</p>
        </div>
    </div>
</body>
</html>
"""

@app.get("/privacy", response_class=HTMLResponse)
async def privacy_policy():
    """Serve the Privacy Policy page"""
    return HTMLResponse(content=PRIVACY_POLICY_HTML)

@app.get("/privacy-policy", response_class=HTMLResponse)
async def privacy_policy_alt():
    """Alternative URL for Privacy Policy"""
    return HTMLResponse(content=PRIVACY_POLICY_HTML)

@app.get("/terms", response_class=HTMLResponse)
async def terms_of_use():
    """Serve the Terms of Use page"""
    return HTMLResponse(content=TERMS_OF_USE_HTML)

@app.get("/terms-of-use", response_class=HTMLResponse)
async def terms_of_use_alt():
    """Alternative URL for Terms of Use"""
    return HTMLResponse(content=TERMS_OF_USE_HTML)

# ======================== ROOT ========================

@api_router.get("/")
async def root():
    return {
        "message": "AchaServiço API",
        "version": "1.0.0",
        "city": "Três Lagoas - MS",
        "payment_method": "pix_manual"
    }

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
